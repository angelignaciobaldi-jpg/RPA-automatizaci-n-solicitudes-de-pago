# -*- coding: utf-8 -*-
"""
RPA - Solicitudes de Pago (SIPP - Petroil)
==========================================
Automatiza el registro de solicitudes de pago a ex-colaboradores (tipo
"Pago Extraordinario" / beneficiario "Acreedor") leyendo el archivo
"BASE DE DATOS LISTADO 1.csv".

DISEÑADO PARA PROBARSE DE FORMA SEGURA:
  - Abre el navegador visible.
  - Procesa solo los registros indicados en config.MAX_REGISTROS.
  - Por defecto SE DETIENE antes de guardar (config.PAUSAR_ANTES_DE_GUARDAR)
    para que revises el formulario a mano.

Uso:
    pip install -r requirements.txt
    playwright install chromium
    python sipp_rpa.py
"""

import csv
import getpass
import logging
import os
import re
import sys
import unicodedata
from datetime import datetime

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

import config
import mapeos


# --------------------------------------------------------------------------- #
#  LOGGING
# --------------------------------------------------------------------------- #
def configurar_log():
    os.makedirs(config.CARPETA_LOGS, exist_ok=True)
    archivo = os.path.join(
        config.CARPETA_LOGS,
        f"rpa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log",
    )
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-7s | %(message)s",
        handlers=[
            logging.FileHandler(archivo, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger("rpa")


log = configurar_log()


# --------------------------------------------------------------------------- #
#  UTILIDADES DE DATOS
# --------------------------------------------------------------------------- #
def normalizar(texto):
    """Quita acentos, espacios sobrantes y pasa a mayúsculas (para comparar)."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.strip().upper()


def leer_csv(ruta):
    """Lee el CSV y devuelve una lista de diccionarios con claves normalizadas."""
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el CSV: {ruta}")

    # utf-8-sig por si el archivo trae BOM (común en Excel).
    with open(ruta, "r", encoding="utf-8-sig", newline="") as f:
        lector = csv.DictReader(f)
        filas = []
        for cruda in lector:
            fila = {normalizar(k): (v.strip() if v else "") for k, v in cruda.items()}
            # Ignora filas totalmente vacías.
            if any(fila.values()):
                filas.append(fila)
    return filas


def _celda_a_texto(v):
    """Convierte el valor de una celda de Excel a texto, igual que vendría del
    CSV (fechas dd/mm/aaaa, números sin notación científica ni '.0')."""
    import datetime as _dt
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Sí" if v else "No"
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def leer_excel(ruta):
    """Lee un Excel (.xlsx) y devuelve la MISMA estructura que leer_csv:
    lista de diccionarios con claves normalizadas y valores en texto."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "Esta versión del programa no incluye soporte de Excel.\n"
            "Actualiza/reinstala el programa, o guarda tu archivo como CSV.")
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el Excel: {ruta}")

    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb.active
    filas, encabezados = [], None
    for celdas in ws.iter_rows(values_only=True):
        valores = [_celda_a_texto(c) for c in celdas]
        if encabezados is None:
            if any(v.strip() for v in valores):       # 1ª fila no vacía = encabezados
                encabezados = [normalizar(v) for v in valores]
            continue
        if not any(v.strip() for v in valores):       # ignora filas vacías
            continue
        fila = {k: v.strip() for k, v in zip(encabezados, valores) if k}
        if any(fila.values()):
            filas.append(fila)
    wb.close()
    return filas


def leer_base(ruta):
    """Lee la base de datos eligiendo el lector según la extensión
    (.csv -> leer_csv, .xlsx/.xls/.xlsm -> leer_excel)."""
    ext = os.path.splitext(ruta)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return leer_excel(ruta)
    return leer_csv(ruta)


def campo(fila, *nombres):
    """Devuelve el primer campo que exista (comparando normalizado)."""
    for n in nombres:
        clave = normalizar(n)
        if clave in fila:
            return fila[clave]
    return ""


def campo_monto(fila):
    """Devuelve el importe a pagar. Acepta varios nombres de columna
    ('MONTO', 'MONTO A PAGAR', 'IMPORTE', etc.) y, si no, cualquier columna
    cuyo nombre contenga 'MONTO' o 'IMPORTE'."""
    v = campo(fila, "MONTO", "MONTO A PAGAR", "IMPORTE", "IMPORTE A PAGAR")
    if v:
        return v
    for clave, valor in fila.items():
        if ("MONTO" in clave or "IMPORTE" in clave) and valor:
            return valor
    return ""


def limpiar_monto(valor):
    """Quita '$', comas y espacios de un importe -> número plano ('1000.00')."""
    return (valor or "").replace("$", "").replace(",", "").strip()


def resolver_descripcion(texto, nombre):
    """Sustituye el texto literal 'NOMBRE DE COLABORADOR' por el nombre real."""
    return re.sub(r"NOMBRE DE(?:L)? COLABORADOR", nombre, texto, flags=re.IGNORECASE)


def _coincide_archivo(ruta, objetivo):
    """Si el archivo 'ruta' corresponde al colaborador 'objetivo' (normalizado),
    devuelve su prioridad por extensión; si no, None."""
    base, ext = os.path.splitext(os.path.basename(ruta))
    if ext.lower() not in config.EXT_CARATULA:
        return None
    nbase = normalizar(base)
    # Coincidencia exacta o por contención (por si el archivo trae sufijos).
    if nbase == objetivo or objetivo in nbase or nbase in objetivo:
        return config.EXT_CARATULA.index(ext.lower())
    return None


def buscar_en(carpeta, archivos, nombre_colaborador):
    """Busca un archivo (pdf/imagen) cuyo nombre coincida con el colaborador,
    tanto en 'carpeta' (si existe) como en la lista 'archivos' (rutas sueltas
    que el operador adjuntó). Devuelve la ruta o None."""
    objetivo = normalizar(nombre_colaborador)
    rutas = list(archivos or [])
    if carpeta and os.path.isdir(carpeta):
        rutas += [os.path.join(carpeta, n) for n in os.listdir(carpeta)]

    candidatos = []
    for ruta in rutas:
        prioridad = _coincide_archivo(ruta, objetivo)
        if prioridad is not None and os.path.isfile(ruta):
            candidatos.append((prioridad, ruta))
    if not candidatos:
        return None
    candidatos.sort()
    return candidatos[0][1]


def buscar_archivo_por_nombre(carpeta, nombre_colaborador):
    """(Compatibilidad) Busca solo en una carpeta."""
    return buscar_en(carpeta, None, nombre_colaborador)


def buscar_caratula(nombre_colaborador):
    """Carátula bancaria: carpeta CARATULAS y/o archivos sueltos."""
    return buscar_en(config.CARPETA_CARATULAS, config.ARCHIVOS_CARATULAS,
                     nombre_colaborador)


def buscar_vobo(nombre_colaborador):
    """Documento Vo.Bo. de Compras: carpeta VOBO y/o archivos sueltos."""
    return buscar_en(config.CARPETA_VOBO, config.ARCHIVOS_VOBO,
                     nombre_colaborador)


# --------------------------------------------------------------------------- #
#  HELPERS DE INTERFAZ (Playwright + plugin "chosen" de SIPP)
# --------------------------------------------------------------------------- #
def seleccionar_chosen(page, select_locator, texto, descripcion, intentos=4):
    """Selecciona una opción en una lista 'chosen' de SIPP.

    El <select> nativo está oculto; junto a él hay un widget 'chosen'. Abrimos
    el widget, escribimos el texto en su buscador y damos clic en el resultado.

    Robustez (equipos lentos / listas dependientes como 'Tipo de Pago
    Extraordinario', que se llena DESPUÉS de elegir el campo padre):
      1) Espera a que el <select> ya tenga la opción buscada (opciones cargadas).
      2) Reintenta varias veces: si el desplegable aún no muestra la opción,
         lo cierra (Escape), espera y lo vuelve a abrir.
    """
    log.info("   - %s = '%s'", descripcion, texto)
    select_locator.wait_for(state="attached", timeout=config.TIMEOUT_MS)

    # 1) Espera a que la opción exista en el <select> (aunque esté oculto).
    try:
        select_locator.locator("option", has_text=texto).first.wait_for(
            state="attached", timeout=config.TIMEOUT_MS)
    except PWTimeout:
        pass  # algunas listas no traen <option> precargada; seguimos igual

    contenedor = select_locator.locator(
        "xpath=following-sibling::div[contains(@class,'chosen-container')][1]")
    contenedor.wait_for(state="visible", timeout=config.TIMEOUT_MS)

    ultimo = None
    for intento in range(1, intentos + 1):
        try:
            contenedor.scroll_into_view_if_needed()
            contenedor.click()
            buscador = contenedor.locator("input.chosen-search-input")
            buscador.fill(texto)
            opcion = contenedor.locator("li.active-result", has_text=texto).first
            opcion.wait_for(state="visible", timeout=6000)
            opcion.click()
            return
        except Exception as e:
            ultimo = e
            log.warning("   (reintento %d/%d en '%s'…)", intento, intentos,
                        descripcion)
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            page.wait_for_timeout(1200)
    raise RuntimeError(
        f"No se pudo seleccionar '{texto}' en '{descripcion}' "
        f"(la lista no cargó la opción a tiempo): {ultimo}")


def cerrar_alertas(page):
    """Cierra cualquier alerta emergente (red-alert) dando clic en su botón
    'Aceptar'. OJO: estos avisos pueden traer también un botón 'Ver detalle';
    hay que clicar específicamente 'Aceptar', no el primero, o se queda en bucle.
    """
    try:
        for _ in range(6):
            btn = page.locator(
                "red-alert button:has-text('Aceptar'):visible").first
            if btn.count() == 0:
                break
            log.info("   (cerrando alerta emergente: Aceptar)")
            btn.click(timeout=5000)
            page.wait_for_timeout(500)
    except Exception:
        pass


def aviso_sin_conceptos(page):
    """True si está visible un aviso de SIPP indicando que la empresa NO tiene
    conceptos de gastos/pago asignados (p.ej. para Acreedores/Deudores)."""
    try:
        alertas = page.locator("red-alert")
        for i in range(alertas.count()):
            el = alertas.nth(i)
            if el.is_visible():
                t = (el.inner_text() or "").lower()
                if "no tiene conceptos" in t:
                    log.info("   Aviso detectado: %s",
                             " ".join(t.split())[:120])
                    return True
    except Exception:
        pass
    return False


def llenar(page, locator, texto, descripcion):
    """Escribe en un input/textarea y dispara los eventos que Angular escucha."""
    log.info("   - %s = '%s'", descripcion, texto)
    locator.wait_for(state="visible", timeout=config.TIMEOUT_MS)
    locator.scroll_into_view_if_needed()
    locator.fill("")
    locator.type(str(texto), delay=15)
    # Forzar el evento change/input de AngularJS.
    locator.dispatch_event("input")
    locator.dispatch_event("change")


# --------------------------------------------------------------------------- #
#  PASOS DEL FLUJO
# --------------------------------------------------------------------------- #
def pedir_credenciales():
    usuario = config.USUARIO or input("Usuario SIPP: ").strip()
    contrasena = config.CONTRASENA or getpass.getpass("Contraseña SIPP: ")
    return usuario, contrasena


def login(page, usuario, contrasena):
    log.info("Abriendo página de login...")
    page.goto(config.URL_LOGIN, timeout=config.TIMEOUT_MS)
    page.locator("#nb_Usuario").fill(usuario)
    page.locator("input[ng-model='de_password']").fill(contrasena)
    page.locator("#btnLogin").click()
    log.info("Login enviado. Esperando a que cargue el sistema...")
    # Espera la redirección fuera de login.html hacia el sistema (index.cfm).
    page.wait_for_url("**/index.cfm*", timeout=config.TIMEOUT_MS)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PWTimeout:
        pass
    log.info("Sistema cargado: %s", page.url)


def configurar_sesion(page, empresa_corta, sucursal):
    """Llena la pantalla 'Configuración de Sesión' (Empresa + Plaza) y Guarda.

    En esa pantalla las empresas se muestran con su nombre largo y el nombre
    corto entre paréntesis, ej: 'ADMINISTRACION DE SERVICIOS ASKE - (Aske )'.
    Por eso buscamos por el nombre corto entre paréntesis: '(Aske )'.
    Devuelve True si la configuró, False si esa pantalla no apareció.
    """
    # Da tiempo a que la SPA (index.cfm) termine de inicializar.
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except PWTimeout:
        pass
    sel_emp = page.locator("select[ng-model='id_Empresa']")
    try:
        sel_emp.wait_for(state="attached", timeout=config.TIMEOUT_MS)
    except PWTimeout:
        log.info("No apareció 'Configuración de Sesión'; se omite ese paso.")
        return False

    log.info("Configurando sesión: empresa='%s', sucursal='%s'",
             empresa_corta, sucursal)
    seleccionar_chosen(page, sel_emp, f"({empresa_corta} )", "Empresa (sesión)")
    # La lista de sucursales se carga al elegir empresa: damos un respiro.
    page.wait_for_timeout(1500)
    seleccionar_chosen(
        page, page.locator("select[ng-model='id_Sucursal']"),
        sucursal, "Plaza/Sucursal (sesión)")
    page.locator("button[ng-click='Guardar()']").click()
    page.wait_for_timeout(1500)
    log.info("Sesión configurada.")
    return True


def ir_a_solicitud_pago(page):
    """Navega a la vista 'Solicitud de Pago - Listado' (ruta #/SolicitudPago)."""
    log.info("Navegando a 'Solicitud de Pago'...")
    try:
        page.evaluate("window.location.hash = '#/SolicitudPago'")
    except Exception:
        # Respaldo: clic directo en el enlace del menú.
        page.locator("a[href='#/SolicitudPago']").first.click()

    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except PWTimeout:
        pass
    # La vista de listado muestra el botón 'Crear Solicitud Pago'.
    page.locator("button[title='Crear Solicitud Pago']").wait_for(
        state="visible", timeout=config.TIMEOUT_MS)
    log.info("En 'Solicitud de Pago - Listado'.")


def asegurar_modo_agregar(page):
    """Se asegura de estar en la vista de 'Agregar' (no en el listado).

    En SIPP el botón con título 'Crear Solicitud Pago' cambia a la vista de
    captura. Si no aparece, asumimos que ya estamos en la vista de agregar.
    """
    try:
        boton_crear = page.locator("button[title='Crear Solicitud Pago']")
        if boton_crear.is_visible():
            log.info("Entrando a la vista de 'Agregar'...")
            boton_crear.click()
    except PWTimeout:
        pass
    # Confirma que está visible el selector de Empresa del formulario.
    page.locator("select[ng-model='solicitudPago.ID_EMPRESA']").wait_for(
        state="attached", timeout=config.TIMEOUT_MS
    )


def llenar_cuenta_bancaria(page, fila, panel_acreedor, correo=""):
    """Abre el modal 'Cuenta Bancaria' y lo llena con los datos del CSV."""
    log.info("   Abriendo modal de Cuenta Bancaria...")
    panel_acreedor.locator("#btn_CuentaBancaria").click()

    modal = page.locator("#divBloqueo_modalCuentaBancaria")
    modal.wait_for(state="visible", timeout=config.TIMEOUT_MS)

    banco = mapeos.traducir(mapeos.BANCOS, campo(fila, "BANCOS"), "BANCOS")
    moneda = mapeos.traducir(mapeos.MONEDAS, campo(fila, "MONEDA"), "MONEDA")
    # Por indicación del usuario, TODAS las transferencias se tratan como SPEI
    # (no se usa 'Mismo Banco' ni el número de cuenta).
    transf = "SPEI"
    nombre_cuenta = campo(fila, "NOMBRE DE CUENTA")
    clabe = campo(fila, "CLAVE INTERBANCARIA")

    seleccionar_chosen(page, page.locator("#id_BancoCB"), banco, "Banco")
    llenar(page, page.locator("#nb_CuentaCB"), nombre_cuenta, "Nombre Cuenta")
    seleccionar_chosen(page, page.locator("#id_MonedaCB"), moneda, "Moneda")
    llenar(page, page.locator("#de_ClabeCB"), clabe, "CLABE Bancaria")
    seleccionar_chosen(
        page, page.locator("#id_TipoTransfCB"), transf, "Tipo Transferencia"
    )

    # Correo (suele ser obligatorio al registrar una cuenta nueva).
    campo_correo = page.locator("#de_EmailCB")
    if correo and campo_correo.is_visible():
        llenar(page, campo_correo, correo, "Correo (cuenta bancaria)")

    # Carátula (archivo obligatorio).
    caratula = buscar_caratula(campo(fila, "EX-COLABORADOR (DESCRIPCIÓN)",
                                     "NOMBRE DE CUENTA"))
    if caratula:
        log.info("   - Carátula = '%s'", caratula)
        subir_con_reintento(page, page.locator("#ar_CaratulaCB"),
                            caratula, "Carátula")
    else:
        log.warning("   - SIN carátula: no se encontró archivo para '%s' en %s",
                    campo(fila, "EX-COLABORADOR (DESCRIPCIÓN)"),
                    config.CARPETA_CARATULAS)

    # Cerrar SIEMPRE el modal: su 'Guardar' (closeModal) solo mete los datos al
    # formulario, NO guarda la solicitud en el servidor. Dejarlo abierto impide
    # que se registren la Moneda y los Conceptos. El guardado real es al final.
    page.locator(
        "#divBloqueo_modalCuentaBancaria button:has-text('Guardar')"
    ).click()
    # Espera a que cierre (la subida de carátula puede tardar). Si aparece el
    # error de Google, lo cierra (clic en 'Aceptar') y deja que la recuperación
    # maneje el registro; así no se queda atorado.
    cerrado = False
    for _ in range(60):  # hasta ~30s
        if not modal.is_visible():
            cerrado = True
            log.info("   Modal de Cuenta Bancaria cerrado.")
            break
        if hay_error_sistema(page):
            log.warning("   Error del sistema (Google) al subir carátula; "
                        "se cierra el aviso.")
            cerrar_alertas(page)
            break
        page.wait_for_timeout(500)
    if not cerrado:
        cerrar_alertas(page)


class AcreedorExistente(Exception):
    """El acreedor ya está en el catálogo: requiere selección manual."""


class SinConceptoPago(Exception):
    """La empresa no tiene asignado el concepto de pago requerido: la solicitud
    se cancela porque no se podría enviar."""


def _cerrar_modal_busqueda(page, modal):
    try:
        modal.locator("button:has-text('Cerrar')").click()
        page.wait_for_timeout(800)
    except Exception:
        pass


def buscar_y_seleccionar_acreedor(page, nombre):
    """Busca el acreedor por nombre. Devuelve:
      - 'nuevo'     : no existe (cierra el buscador).
      - 'existente' : existe exactamente uno y lo SELECCIONA (doble clic).
      - 'ambiguo'   : hay varias coincidencias (cierra; requiere revisión).
    """
    log.info("   Buscando si el acreedor '%s' ya existe...", nombre)
    cerrar_alertas(page)
    page.locator("a[ng-click='ayudaProveedores(2)']:visible").first.click()
    modal = page.locator("#divBloqueo_modalProveedores")
    modal.wait_for(state="visible", timeout=config.TIMEOUT_MS)
    modal.locator("input[ng-model='nb_proveedor']").fill(nombre)
    modal.locator("button[ng-click='BuscarProveedores()']").click()
    page.wait_for_timeout(2500)
    n = modal.locator(".ngRow").count()
    log.info("   -> %d resultado(s) en el catálogo de acreedores.", n)
    if n == 0:
        _cerrar_modal_busqueda(page, modal)
        return "nuevo"
    if n > 1:
        log.warning("   %d coincidencias para '%s' (ambiguo).", n, nombre)
        _cerrar_modal_busqueda(page, modal)
        return "ambiguo"
    # Exactamente una: seleccionar con doble clic.
    modal.locator(".ngRow").first.dblclick()
    page.wait_for_timeout(2000)
    cerrar_alertas(page)
    return "existente"


def seleccionar_forma_y_tipo(page):
    """Selecciona Forma de Pago y Tipo de Gasto (común a ambos caminos)."""
    if config.FORMA_DE_PAGO:
        seleccionar_chosen(page, page.locator("#id_MetodoPago"),
                           config.FORMA_DE_PAGO, "Forma de Pago")
    if config.TIPO_DE_GASTO:
        seleccionar_chosen(page, page.locator("#id_TipoGastoDdc"),
                           config.TIPO_DE_GASTO, "Tipo de Gasto")


def seleccionar_cuenta_existente(page, fila):
    """Selecciona la cuenta bancaria existente del acreedor en 'Cuentas Ban',
    emparejando por la CLABE del CSV (la opción se muestra como 'CLABE - NOMBRE')."""
    clabe = campo(fila, "CLAVE INTERBANCARIA")
    seleccionar_chosen(page, page.locator("#ID_CUENTABANCARIAPROVEEDOR"),
                       clabe, "Cuenta bancaria existente")


def _fila_concepto(page, concepto):
    """Devuelve el renglón del grid de Conceptos cuyo texto contiene TODAS las
    palabras del 'concepto' (sin distinguir mayúsculas/acentos). None si no hay.

    Ej.: concepto 'PAGO PTU' encuentra cualquier renglón que tenga 'PAGO' y
    'PTU' (sirve aunque diga 'PAGO PTU', 'PAGO DE PTU', etc.).
    """
    palabras = [normalizar(w) for w in concepto.split() if w.strip()]
    grid = page.locator("#id_ConceptosPagoGrid .ngRow")
    for i in range(grid.count()):
        fila = grid.nth(i)
        try:
            texto = normalizar(fila.locator(".ngCellText").first.inner_text())
        except Exception:
            continue
        if palabras and all(w in texto for w in palabras):
            return fila
    return None


def concepto_disponible(page, concepto=None):
    """Revisa (temprano) si la empresa tiene el concepto de pago en el grid.
    - True  si hay un renglón que coincide (todas las palabras del concepto).
    - False si el grid quedó VACÍO o ningún renglón coincide.
    - None  solo si no se pudo abrir la pestaña.
    """
    concepto = concepto or config.CONCEPTO_PAGO
    try:
        page.locator("li a", has_text="Conceptos de Pago").first.click()
    except Exception:
        return None
    grid = page.locator("#id_ConceptosPagoGrid .ngRow")
    # Espera hasta ~12s a que aparezcan renglones (rompe en cuanto haya).
    for _ in range(24):
        if grid.count() > 0:
            break
        page.wait_for_timeout(500)
    if grid.count() == 0:
        return False  # grid vacío tras esperar = empresa sin conceptos
    return _fila_concepto(page, concepto) is not None


def llenar_concepto_pago(page, monto, concepto=None):
    """Activa la pestaña 'Conceptos de Pago' y captura el importe en el renglón
    cuyo texto contiene las palabras del concepto (marca su 'Seleccionar')."""
    concepto = concepto or config.CONCEPTO_PAGO
    log.info("   Conceptos de Pago: buscar '%s' = %s", concepto, monto)
    # Activa la pestaña (encabezado tipo Bootstrap).
    page.locator("li a", has_text="Conceptos de Pago").first.click()
    page.wait_for_timeout(1000)
    grid = page.locator("#id_ConceptosPagoGrid .ngRow")
    for _ in range(16):
        if grid.count() > 0:
            break
        page.wait_for_timeout(500)
    fila = _fila_concepto(page, concepto)
    if fila is None:
        total = grid.count()
        raise SinConceptoPago(
            f"La empresa no tiene un concepto con las palabras '{concepto}' "
            f"({total} concepto(s) disponibles). Se cancela la solicitud.")
    try:
        log.info("   Concepto encontrado: %s",
                 fila.locator(".ngCellText").first.inner_text().strip())
    except Exception:
        pass
    fila.scroll_into_view_if_needed()
    monto_input = fila.locator("input[ng-model='row.entity.IM_IMPORTE']").first
    # 1) Captura el monto TECLEÁNDOLO (la directiva de moneda 'contenido_moneda'
    #    necesita pulsaciones reales; con fill() el modelo no se confirma bien).
    monto_input.click()
    try:
        monto_input.press("Control+a")
    except Exception:
        pass
    monto_input.type(str(monto), delay=30)
    monto_input.press("Tab")          # blur: confirma el importe en el modelo
    page.wait_for_timeout(400)
    # 2) Selecciona el renglón con la CELDA de selección (div.ngSelectionCell).
    #    Esto agrega el concepto a 'ar_ConceptosGastosSelc' y dispara el recálculo
    #    del total (im_TotalConceptos / IM_CANTIDADPAGAR). El total suma SOLO los
    #    conceptos seleccionados, por eso esto es lo que evita el $0.
    #    OJO: NO clicar el nombre u otra celda de la fila -> en ng-grid eso
    #    DESELECCIONA el renglón.
    ya_sel = False
    try:
        ya_sel = bool(monto_input.evaluate(
            "el => !!angular.element(el).scope().row.selected"))
    except Exception:
        pass
    if not ya_sel:
        fila.locator("div.ngSelectionCell").first.click()
        page.wait_for_timeout(500)
    # 3) Respaldo: si el total quedó en 0, fuerza el recálculo por AngularJS.
    try:
        total = _leer_total(page)
        if _es_cero(total):
            monto_input.evaluate(
                "el => { const s = angular.element(el).scope();"
                " if (typeof s.cambioImporteConceptos === 'function') {"
                " s.cambioImporteConceptos(s.row); s.$root.$apply(); } }")
            page.wait_for_timeout(300)
            total = _leer_total(page)
        log.info("   Cantidad a Pagar (total): %s", total or "(no leído)")
        if _es_cero(total):
            log.warning("   ¡OJO! El total sigue en 0 tras capturar el concepto.")
    except Exception:
        pass


def _es_cero(total):
    return (total or "0").replace("$", "").replace(",", "").strip() in (
        "", "0", "0.00", ".00")


def _leer_total(page):
    """Devuelve el texto del total de la solicitud (Cantidad a Pagar), si se
    puede leer; '' si no. Solo informativo para la bitácora."""
    for sel in ("input[ng-model='im_TotalConceptos']",
                "input[ng-model='solicitudPago.IM_TOTAL']",
                "input[ng-model='solicitudPago.IM_CANTIDADPAGAR']"):
        loc = page.locator(sel)
        try:
            if loc.count() > 0:
                v = loc.first.input_value()
                if v:
                    return v
        except Exception:
            continue
    return ""


def guardar_solicitud(page):
    """Da clic en 'Guardar' (generarSolicitud) y confirma el diálogo."""
    log.info("   Guardando solicitud...")
    page.locator("button[ng-click='generarSolicitud()']").click()
    # Diálogo: "¿Desea guardar la solicitud de pago?" -> Aceptar (si aparece).
    aceptar = page.locator("#__btn_aceptarConfirm__")
    try:
        aceptar.wait_for(state="visible", timeout=config.TIMEOUT_MS)
        log.info("   Confirmando guardado (Aceptar)...")
        aceptar.click()
    except PWTimeout:
        log.info("   (sin diálogo de confirmación)")
    page.wait_for_timeout(4000)
    log.info("   Solicitud guardada.")


def adjuntar_vobo(page, nombre):
    """Tras guardar la solicitud, adjunta el Vo.Bo. de Compras en la pestaña
    'Documentos Respaldo' (si existe el archivo para el colaborador)."""
    vobo = buscar_vobo(nombre)
    if not vobo:
        log.info("   Sin Vo.Bo. para '%s' (no se adjunta).", nombre)
        return
    log.info("   Adjuntando Vo.Bo.: %s", vobo)
    cerrar_alertas(page)
    page.locator("li a", has_text="Documentos Respaldo").first.click()
    page.wait_for_timeout(1000)
    # El '+' agrega un renglón con un input de archivo.
    page.locator("a[ng-click='agregarDocumento()']").first.click()
    page.wait_for_timeout(1000)
    entrada = page.locator("input.fileGrid[type='file']").last
    subir_con_reintento(page, entrada, vobo, "Vo.Bo.")
    cerrar_alertas(page)
    log.info("   Vo.Bo. adjuntado.")


def hay_error_sistema(page):
    """True si hay un aviso de error del sistema visible, incluido el error de
    Google ('Error making Google REST call' / CloudStorage 403 Forbidden)."""
    try:
        al = page.locator("red-alert")
        for i in range(al.count()):
            el = al.nth(i)
            if not el.is_visible():
                continue
            t = (el.inner_text() or "").lower()
            if ("error del sistema" in t or "google rest" in t
                    or "making google" in t or "cloudstorage" in t
                    or "403 forbidden" in t):
                return True
    except Exception:
        pass
    return False


def subir_con_reintento(page, file_input, ruta, etiqueta, intentos=3):
    """Sube un archivo y reintenta si aparece 'Error del Sistema' (subida a
    Google intermitente). Lanza RuntimeError si falla tras los reintentos."""
    for intento in range(1, intentos + 1):
        file_input.set_input_files(ruta)
        page.wait_for_timeout(4000)  # da tiempo a la subida
        if hay_error_sistema(page):
            log.warning("   Subida de %s falló (intento %d/%d). Reintentando...",
                        etiqueta, intento, intentos)
            cerrar_alertas(page)
            page.wait_for_timeout(1500)
            continue
        return  # éxito
    raise RuntimeError(f"No se pudo subir {etiqueta} tras {intentos} intentos "
                       f"(error de subida a Google).")


def solicitar_autorizacion(page):
    """Da clic en 'Solicitar Autorización' (enviarAutorizacion) y confirma."""
    if not config.SOLICITAR_AUTORIZACION:
        return
    log.info("   Solicitando autorización...")
    cerrar_alertas(page)
    boton = page.locator("button[ng-click='enviarAutorizacion()']")
    boton.wait_for(state="visible", timeout=config.TIMEOUT_MS)
    boton.click()
    page.wait_for_timeout(1500)
    # Posible diálogo de confirmación.
    try:
        aceptar = page.locator("#__btn_aceptarConfirm__")
        if aceptar.count() > 0 and aceptar.is_visible():
            aceptar.click()
    except Exception:
        pass
    page.wait_for_timeout(3000)
    cerrar_alertas(page)
    log.info("   Autorización solicitada.")


def reset_a_agregar(page):
    """Deja listo un formulario de Agregar limpio para el siguiente registro.

    Si algo quedó atorado (p.ej. el error de Google con un modal abierto que no
    cierra), RECARGA la página para resetear y vuelve a navegar al listado.
    """
    cerrar_alertas(page)
    # 1) Intento rápido y normal: 'Regresar al Listado'.
    try:
        btn = page.locator("button[title='Regresar al Listado']")
        if btn.count() > 0 and btn.first.is_visible():
            btn.first.click(timeout=5000)
            page.wait_for_timeout(1200)
            cerrar_alertas(page)
            if page.locator("button[title='Crear Solicitud Pago']").first.is_visible():
                return
    except Exception:
        pass

    # 2) Recuperación DURA: recargar la página y re-navegar.
    log.warning("   Recuperando proceso: recargando la página...")
    try:
        page.reload(timeout=config.TIMEOUT_MS)
    except Exception:
        try:
            page.goto(config.URL_LOGIN.replace("login.html", "index.cfm"),
                      timeout=config.TIMEOUT_MS)
        except Exception:
            pass
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PWTimeout:
        pass
    cerrar_alertas(page)
    # Si reapareció 'Configuración de Sesión', reconfigura.
    if page.locator("select[ng-model='id_Empresa']").count() > 0:
        emp = mapeos.traducir(
            mapeos.EMPRESAS, config.EMPRESA_SESION.strip() or "ASKE", "EMPRESA")
        configurar_sesion(page, emp, config.SUCURSAL_SESION)
    ir_a_solicitud_pago(page)
    cerrar_alertas(page)


def llenar_solicitud(page, fila):
    """Llena UN registro completo en el formulario de Agregar."""
    nombre = campo(fila, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA")
    log.info("Llenando solicitud para: %s", nombre)

    asegurar_modo_agregar(page)

    empresa = mapeos.traducir(mapeos.EMPRESAS, campo(fila, "EMPRESA"), "EMPRESA")
    sucursal = mapeos.traducir(mapeos.SUCURSALES, campo(fila, "SUCURSAL"), "SUCURSAL")

    # 1) Encabezado.
    seleccionar_chosen(
        page, page.locator("select[ng-model='solicitudPago.ID_EMPRESA']"),
        empresa, "Empresa")
    seleccionar_chosen(
        page, page.locator("select[ng-model='solicitudPago.ID_SUCURSAL']"),
        sucursal, "Sucursal")
    seleccionar_chosen(
        page, page.locator("select[ng-model='solicitudPago.ID_TIPOPAGO']"),
        config.TIPO_PAGO, "Tipo de pago")
    seleccionar_chosen(
        page, page.locator("select[ng-model='solicitudPago.ID_TIPOPAGOEXTRAORDINARIO']"),
        config.TIPO_BENEFICIARIO, "Tipo de Pago Extraordinario")

    # Da tiempo a que SIPP muestre el aviso al elegir empresa/beneficiario.
    page.wait_for_timeout(1500)

    # Si SIPP avisa que la empresa NO tiene conceptos asignados (para Acreedores/
    # Deudores), se CANCELA la solicitud: no se podría enviar.
    if aviso_sin_conceptos(page):
        cerrar_alertas(page)
        raise SinConceptoPago(
            "La empresa no tiene conceptos de gastos asignados para Acreedores. "
            "Se cancela la solicitud (no se podría enviar).")

    # Cierra otros avisos benignos que tapan la pantalla.
    cerrar_alertas(page)

    # Respaldo: si en el grid de conceptos no está el concepto requerido, cancela.
    if concepto_disponible(page) is False:
        raise SinConceptoPago(
            f"La empresa no tiene asignado el concepto '{config.CONCEPTO_PAGO}'. "
            f"Se cancela la solicitud (no se podría enviar).")

    # 2) Buscar-o-crear acreedor.
    estado = buscar_y_seleccionar_acreedor(page, nombre)
    if estado == "ambiguo":
        # Varias coincidencias: no arriesgamos elegir mal -> revisión manual.
        raise AcreedorExistente(nombre)

    if estado == "existente":
        # ----- ACREEDOR YA REGISTRADO: solo seleccionar su cuenta -----
        log.info("   Acreedor EXISTENTE seleccionado (no se recaptura banco).")
        seleccionar_forma_y_tipo(page)
        seleccionar_cuenta_existente(page, fila)
    else:
        # ----- ACREEDOR NUEVO: registrar todo -----
        log.info("   Acreedor NUEVO: se registra.")
        if config.BENEFICIARIO_NO_REGISTRADO:
            log.info("   - Marcando 'No Registrado'")
            chk = page.locator("#sn_NoProveedor")
            if not chk.is_checked():
                # El input está cubierto por su <label> estilizado: clic en la etiqueta.
                page.locator("label[for='sn_NoProveedor']").click()

        # Panel del Acreedor (ng-show ...== 3). Acota campos con id duplicado.
        panel_acreedor = page.locator(
            "div.row[ng-show*='ID_TIPOPAGOEXTRAORDINARIO == 3']")
        llenar(page, panel_acreedor.locator("#de_RazonSocial"),
               nombre, "Descripción del Acreedor")
        llenar(page, panel_acreedor.locator("#de_RFC"),
               campo(fila, "RFC"), "RFC")
        correo = (campo(fila, "CORREO", "EMAIL", "CORREO ELECTRONICO")
                  or config.CORREO_ACREEDOR)
        if correo:
            llenar(page, panel_acreedor.locator("#de_EmailReg"),
                   correo, "Correo acreedor")

        seleccionar_forma_y_tipo(page)

        # Cuenta bancaria nueva (modal).
        llenar_cuenta_bancaria(page, fila, panel_acreedor, correo)

    # 5) Carátula en el campo "PDF" del acreedor (requerido en AMBOS casos).
    caratula = buscar_caratula(nombre)
    if caratula:
        log.info("   - Carátula (campo PDF acreedor) = '%s'", caratula)
        subir_con_reintento(page, page.locator("#ar_Pdf"),
                            caratula, "Carátula (PDF)")
    else:
        log.warning("   - SIN carátula para el campo PDF: '%s' en %s",
                    nombre, config.CARPETA_CARATULAS)

    # 6) Fecha y descripción de la solicitud.
    fecha = campo(fila, "FECHA DE PAGO", "FECHA")
    if fecha:
        llenar(page, page.locator("#fh_pago"), fecha, "Fecha de pago")

    descripcion = campo(fila, "DESCRIPCION", "DESCRIPCIÓN SOLICITUD")
    if descripcion:
        descripcion = resolver_descripcion(descripcion, nombre)
        llenar(page, page.locator("textarea[ng-model='solicitudPago.DE_DESCRIPCION']"),
               descripcion, "Descripción")

    # 7) Importe en 'Conceptos de Pago' ('Cantidad a Pagar' está deshabilitada
    #    en Pago Extraordinario).
    monto = limpiar_monto(campo_monto(fila))
    if monto:
        llenar_concepto_pago(page, monto)
    else:
        log.warning("   - MONTO vacío en el CSV para %s", nombre)

    log.info("   Formulario llenado.")


def validar_datos(filas):
    """Valida las filas y devuelve un resumen para la vista previa de la app."""
    hay_car = os.path.isdir(config.CARPETA_CARATULAS) or bool(config.ARCHIVOS_CARATULAS)
    hay_vobo = os.path.isdir(config.CARPETA_VOBO) or bool(config.ARCHIVOS_VOBO)
    total = len(filas)
    con_monto = 0
    problemas = []
    for fila in filas:
        nombre = campo(fila, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA")
        errs = []
        for dic, col, et in [
            (mapeos.EMPRESAS, "EMPRESA", "Empresa"),
            (mapeos.SUCURSALES, "SUCURSAL", "Sucursal"),
            (mapeos.BANCOS, "BANCOS", "Banco"),
            (mapeos.MONEDAS, "MONEDA", "Moneda"),
        ]:
            try:
                mapeos.traducir(dic, campo(fila, col), et)
            except ValueError:
                errs.append(f"{et} no reconocido: '{campo(fila, col)}'")
        clabe = campo(fila, "CLAVE INTERBANCARIA")
        if not (clabe.isdigit() and len(clabe) == 18):
            errs.append(f"CLABE inválida ({len(clabe)} dígitos)")
        monto_txt = campo_monto(fila)
        monto = limpiar_monto(monto_txt)
        if monto:
            con_monto += 1
            try:
                float(monto)
            except ValueError:
                errs.append(f"MONTO no numérico: '{monto_txt}'")
        if hay_car and monto and not buscar_caratula(nombre):
            errs.append("Falta carátula")
        if errs:
            problemas.append((nombre, errs))
    return {
        "total": total,
        "con_monto": con_monto,
        "sin_monto": total - con_monto,
        "hay_caratulas": hay_car,
        "hay_vobo": hay_vobo,
        "problemas": problemas,
    }


# --------------------------------------------------------------------------- #
#  NÚCLEO REUTILIZABLE (lo usa la CLI y la interfaz gráfica)
# --------------------------------------------------------------------------- #
def procesar(usuario, contrasena, filas, on_progreso=None, detener=None,
             escribir_rep=True):
    """Ejecuta el robot sobre 'filas' (ya leídas del CSV).

    - usuario/contrasena: credenciales SIPP.
    - on_progreso(dict): callback opcional para avisar avance. Recibe claves
      i, total, nombre, estado ('procesando'|'ok'|'omitido'|'revisar'|'error').
    - detener(): callback opcional que devuelve True para abortar entre registros.

    Devuelve un resumen: {exitos, errores, revisar, omitidos}.
    No pide credenciales ni usa input() en modo automático (PAUSAR=False).
    """
    def avisar(**kw):
        if on_progreso:
            try:
                on_progreso(kw)
            except Exception:
                pass

    exitos, errores, revisar, omitidos, cancelados = 0, [], [], [], []
    resultados = []  # detalle por registro para el reporte final
    total = len(filas)
    with sync_playwright() as p:
        navegador = p.chromium.launch(headless=not config.NAVEGADOR_VISIBLE)
        contexto = navegador.new_context()
        contexto.set_default_timeout(config.TIMEOUT_MS)
        page = contexto.new_page()
        try:
            login(page, usuario, contrasena)

            # Configuración de sesión + navegación al módulo.
            if config.CONFIGURAR_SESION:
                emp = config.EMPRESA_SESION.strip() or campo(filas[0], "EMPRESA")
                emp = mapeos.traducir(mapeos.EMPRESAS, emp, "EMPRESA_SESION")
                configurar_sesion(page, emp, config.SUCURSAL_SESION)
            ir_a_solicitud_pago(page)

            for i, fila in enumerate(filas, 1):
                if detener and detener():
                    log.warning("Proceso detenido por el usuario.")
                    break
                nombre = campo(fila, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA")
                empresa = campo(fila, "EMPRESA")
                banco = campo(fila, "BANCOS")
                monto_txt = campo_monto(fila).strip()
                log.info("----- Registro %d/%d: %s -----", i, total, nombre)
                avisar(i=i, total=total, nombre=nombre, estado="procesando")

                def anotar(estado, detalle=""):
                    # Guarda el resultado y lo ENVÍA en vivo (para que la app
                    # pueda generar el reporte en cualquier momento).
                    r = {"nombre": nombre, "empresa": empresa, "banco": banco,
                         "monto": monto_txt or "—", "estado": estado,
                         "detalle": detalle}
                    resultados.append(r)
                    avisar(i=i, total=total, nombre=nombre,
                           estado=estado.lower(), resultado=r)

                if not limpiar_monto(monto_txt):
                    log.info("   SIN MONTO: se omite.")
                    omitidos.append(nombre)
                    anotar("OMITIDO", "Sin monto")
                    continue
                try:
                    llenar_solicitud(page, fila)
                    if config.PAUSAR_ANTES_DE_GUARDAR:
                        input(">>> Revisa el formulario y presiona ENTER...")
                    else:
                        guardar_solicitud(page)
                        adjuntar_vobo(page, nombre)
                        solicitar_autorizacion(page)
                    exitos += 1
                    log.info("   OK: %s procesado.", nombre)
                    anotar("OK")
                except AcreedorExistente:
                    log.warning("REVISAR: '%s' tiene varias coincidencias "
                                "(ambiguo); no se elige automáticamente.", nombre)
                    revisar.append(nombre)
                    anotar("REVISAR", "Acreedor con varias coincidencias (ambiguo)")
                except SinConceptoPago as e:
                    log.warning("CANCELADO (%s): %s", nombre, e)
                    cancelados.append(nombre)
                    anotar("CANCELADO", str(e))
                except Exception as e:
                    log.error("ERROR en registro %d (%s): %s", i, nombre, e)
                    errores.append((nombre, str(e)))
                    try:
                        page.screenshot(path=os.path.join(
                            config.CARPETA_LOGS,
                            f"error_{i}_{datetime.now():%H%M%S}.png"))
                    except Exception:
                        pass
                    anotar("ERROR", descripcion_error(str(e)))
                # Reset del formulario para el siguiente registro.
                if i < total:
                    try:
                        reset_a_agregar(page)
                    except Exception:
                        pass
        finally:
            navegador.close()

    log.info("=== FIN. Éxitos: %d | Errores: %d | Cancelados: %d | "
             "Revisar: %d | Omitidos: %d ===", exitos, len(errores),
             len(cancelados), len(revisar), len(omitidos))
    ruta_reporte = ""
    if escribir_rep:
        ruta_reporte = escribir_reporte(resultados)
        if ruta_reporte:
            log.info("Reporte final: %s", ruta_reporte)
    return {"exitos": exitos, "errores": errores, "revisar": revisar,
            "omitidos": omitidos, "cancelados": cancelados,
            "reporte": ruta_reporte, "resultados": resultados,
            "detenido": bool(detener and detener())}


def descripcion_error(texto):
    """Traduce un error técnico (Playwright, etc.) a una nota clara para el
    usuario, que le dé una idea de qué pasó."""
    t = (texto or "").lower()
    if "set_input_files" in t:
        return ("No se pudo adjuntar un archivo (carátula o Vo.Bo.): el campo "
                "no estuvo disponible a tiempo.")
    if "google rest" in t or "cloudstorage" in t or "403" in t:
        return ("Error del sistema al subir el archivo al almacenamiento "
                "(servicio no disponible).")
    if "no hay mapeo" in t or "no reconocido" in t:
        return texto  # ya es claro (banco/empresa/moneda no reconocido)
    if "select_option" in t or "chosen" in t or "li.active-result" in t:
        return ("No se pudo seleccionar una opción de una lista (banco, empresa, "
                "etc.): no apareció a tiempo.")
    if ".click" in t and "timeout" in t:
        return ("No se pudo dar clic en un botón o elemento: la página no "
                "respondió a tiempo.")
    if ("fill" in t or ".type" in t) and "timeout" in t:
        return "No se pudo escribir en un campo: no apareció a tiempo."
    if "wait_for" in t and "timeout" in t:
        return ("Un elemento esperado no apareció a tiempo (la página tardó o "
                "cambió de forma inesperada).")
    if "timeout" in t:
        return "La página tardó demasiado en responder."
    if "navigation" in t or "net::" in t or "err_" in t:
        return "Problema de conexión o de carga de la página."
    # Por defecto: primera línea recortada.
    return ((texto or "Error desconocido").splitlines()[0])[:160]


def escribir_reporte(resultados):
    """Escribe un reporte final (CSV) con el resultado de cada registro en la
    carpeta de logs. Devuelve la ruta del archivo (o '' si no se pudo)."""
    if not resultados:
        return ""
    try:
        os.makedirs(config.CARPETA_LOGS, exist_ok=True)
        ruta = os.path.join(
            config.CARPETA_LOGS,
            f"reporte_{datetime.now():%Y%m%d_%H%M%S}.csv")
        with open(ruta, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Colaborador", "Empresa", "Banco", "Monto",
                        "Estado", "Nota"])
            for r in resultados:
                w.writerow([r["nombre"], r["empresa"], r.get("banco", ""),
                            r["monto"], r["estado"], r["detalle"]])
        return ruta
    except Exception as e:
        log.error("No se pudo escribir el reporte: %s", e)
        return ""


# --------------------------------------------------------------------------- #
#  MAIN (línea de comandos)
# --------------------------------------------------------------------------- #
def main():
    log.info("=== RPA Solicitudes de Pago (SIPP) — %s ===", config.AMBIENTE)
    filas = leer_csv(config.ARCHIVO_CSV)
    log.info("Registros leídos del CSV: %d", len(filas))
    if config.MAX_REGISTROS is not None:
        filas = filas[: config.MAX_REGISTROS]
        log.info("Procesando solo los primeros %d (config.MAX_REGISTROS).",
                 len(filas))
    usuario, contrasena = pedir_credenciales()
    resumen = procesar(usuario, contrasena, filas)
    for nombre, err in resumen["errores"]:
        log.info("   ERROR  - %s: %s", nombre, err)
    for nombre in resumen.get("cancelados", []):
        log.info("   CANCELADO - %s (empresa sin concepto)", nombre)
    for nombre in resumen["revisar"]:
        log.info("   REVISAR - %s (ambiguo)", nombre)
    for nombre in resumen["omitidos"]:
        log.info("   OMITIDO - %s (sin monto)", nombre)


if __name__ == "__main__":
    main()
