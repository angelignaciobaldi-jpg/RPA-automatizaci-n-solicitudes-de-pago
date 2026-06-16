# -*- coding: utf-8 -*-
"""
RPA Solicitudes de Pago (SIPP) — Interfaz gráfica
=================================================
App de escritorio para que un operador: cargue el CSV, vea una vista previa
validada, escriba sus credenciales y ejecute el robot automáticamente.

Apunta SIEMPRE a PRODUCCIÓN. Para correr:
    python app.py
"""

import os
import sys
import queue
import logging
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


def ruta_recurso(rel):
    """Ruta a un recurso (logo, navegador) que funciona corriendo como script y
    como .exe empaquetado con PyInstaller (que usa sys._MEIPASS)."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


# Navegador Chromium: en el .exe va EMPAQUETADO dentro (carpeta 'ms-playwright');
# corriendo como script usa el instalado en el equipo. Debe fijarse ANTES de
# importar sipp_rpa (que importa playwright).
if getattr(sys, "frozen", False):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = ruta_recurso("ms-playwright")

import config
import sipp_rpa
import mapeos

# ===================================================================== #
#  AMBIENTE DEL EJECUTABLE
#  "PRODUCCION" = versión final | "PRUEBAS" = preprod (para probar)
# ===================================================================== #
AMBIENTE_APP = "PRODUCCION"   # PRUEBAS=preprod (validación) | PRODUCCION=final
# ===================================================================== #

config.AMBIENTE = AMBIENTE_APP
config.URL_LOGIN = config.URLS[AMBIENTE_APP]
# En pruebas (preprod) NO existe "PAGO PTU"; se usa uno que sí existe allá.
config.CONCEPTO_PAGO = ("PAGO PTU" if AMBIENTE_APP == "PRODUCCION"
                        else "PAGO DE NOMINA")
config.PAUSAR_ANTES_DE_GUARDAR = False

COLOR_OK = "#1a7f37"
COLOR_MAL = "#bb0000"

# --- Paleta de la interfaz (moderno claro corporativo) ---
UI_FUENTE = "Segoe UI"
UI_BG = "#eef1f5"        # fondo de la app
UI_CARD = "#ffffff"      # fondo de las tarjetas
UI_AZUL = "#00437f"      # azul corporativo (encabezado / acción primaria)
UI_VERDE = "#1a7f37"     # acción de éxito (aplicar / exportar)
UI_ROJO = "#bb0000"      # acción de detener
UI_BORDE = "#d8dee6"     # borde sutil de tarjetas
UI_TXT = "#1f2d3d"       # texto principal
UI_TENUE = "#6b7785"     # texto secundario / ayudas

# Repo para avisar de nuevas versiones (mismo que usa el lanzador).
REPO_GH = "angelignaciobaldi-jpg/RPA-automatizaci-n-solicitudes-de-pago"
RAMA_GH = "main"
API_COMMITS = f"https://api.github.com/repos/{REPO_GH}/commits/{RAMA_GH}"
INTERVALO_UPDATE_MS = 180000   # revisa cada 3 minutos


class HandlerCola(logging.Handler):
    """Manda los mensajes del log a una cola para mostrarlos en la ventana."""
    def __init__(self, cola):
        super().__init__()
        self.cola = cola

    def emit(self, record):
        self.cola.put(("log", self.format(record)))


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"RPA Solicitudes de Pago — SIPP ({config.AMBIENTE})")
        self.geometry("960x720")
        self.minsize(860, 640)

        self.filas = None
        self.ruta_csv = None
        self.cola = queue.Queue()
        self.detener_flag = threading.Event()
        self.worker = None
        # Control de avance / reanudar.
        self.filas_run = []                # registros seleccionados a procesar
        self.fila_check = {}               # id(fila) -> incluido (casilla ✓)
        self.procesados_ok = set()         # claves de registros ya procesados (OK)
        self.siguiente = 0                 # índice del próximo registro a procesar
        self.resultados_global = []        # detalle acumulado entre corridas
        self.usuario = ""
        self.contrasena = ""
        self.estado = "idle"               # idle|corriendo|detenido|fin

        self._construir()
        self._enganchar_log()
        self.after(150, self._drenar_cola)
        self._avisar_actualizacion()
        # Aviso de nueva versión: revisa GitHub periódicamente (en 2º plano).
        self.version_actual = None          # SHA del último commit conocido
        self.aviso_update_visible = False
        self.after(4000, self._chequear_actualizacion)

    def _avisar_actualizacion(self):
        """Muestra en el log si el lanzador actualizó código desde GitHub."""
        upd = os.environ.get("RPA_ACTUALIZADOS")
        if upd is None:
            return
        try:
            k = int(upd)
        except ValueError:
            return
        if k > 0:
            self._log_ui(f"Actualización: se descargaron {k} archivo(s) de "
                         f"código nuevos/actualizados desde GitHub.")
        else:
            self._log_ui("Código verificado con GitHub: ya estaba al día.")

    # ------------------------------------------------------------------ #
    #  Aviso de nueva versión disponible
    # ------------------------------------------------------------------ #
    def _chequear_actualizacion(self):
        """Cada cierto tiempo, revisa en 2º plano si hay una versión nueva."""
        threading.Thread(target=self._hilo_chequeo, daemon=True).start()
        self.after(INTERVALO_UPDATE_MS, self._chequear_actualizacion)

    def _hilo_chequeo(self):
        sha = self._obtener_sha_remoto()
        if sha:
            self.cola.put(("version", sha))

    @staticmethod
    def _obtener_sha_remoto():
        """Devuelve el SHA del último commit en GitHub, o None si no hay red."""
        try:
            import json
            import urllib.request
            req = urllib.request.Request(
                API_COMMITS, headers={"User-Agent": "RPA-SIPP",
                                      "Accept": "application/vnd.github.sha"})
            with urllib.request.urlopen(req, timeout=12) as r:
                cuerpo = r.read().decode().strip()
            # Con Accept sha devuelve el hash plano; por si acaso, soporta JSON.
            if cuerpo.startswith("{"):
                return json.loads(cuerpo).get("sha")
            return cuerpo or None
        except Exception:
            return None

    def _on_version(self, sha):
        """Llega un SHA del repo. Fija la versión base o avisa si cambió."""
        if self.version_actual is None:
            self.version_actual = sha          # versión con la que se abrió
            return
        if sha == self.version_actual or self.aviso_update_visible:
            return
        if self.estado == "corriendo":
            return   # no interrumpir un proceso en curso; avisa al terminar
        self._mostrar_aviso_update(sha)

    def _mostrar_aviso_update(self, sha):
        self.aviso_update_visible = True
        win = tk.Toplevel(self)
        win.title("Actualización disponible")
        win.transient(self)
        win.resizable(False, False)
        try:
            win.grab_set()
        except Exception:
            pass
        tk.Label(win, text="🔄  Hay una nueva versión del programa",
                 font=("Segoe UI", 11, "bold"), fg="#00437f"
                 ).pack(padx=24, pady=(18, 4))
        tk.Label(win, justify="center",
                 text=("Se publicó una actualización.\n"
                       "'Actualizar ahora' reinicia la app para aplicarla.\n"
                       "'Enterado' cierra este aviso y sigues trabajando.")
                 ).pack(padx=24, pady=(0, 14))

        def actualizar():
            win.destroy()
            self._actualizar_ahora()

        def enterado():
            # No volver a avisar por ESTA versión; una futura sí avisará.
            self.version_actual = sha
            self.aviso_update_visible = False
            win.destroy()

        btns = tk.Frame(win)
        btns.pack(pady=(0, 16))
        tk.Button(btns, text="🔄  Actualizar ahora", bg="#00437f", fg="white",
                  font=("Segoe UI", 10, "bold"), width=18,
                  command=actualizar).pack(side="left", padx=8)
        tk.Button(btns, text="Enterado", width=12,
                  command=enterado).pack(side="left", padx=8)
        win.protocol("WM_DELETE_WINDOW", enterado)
        win.update_idletasks()
        # Centra sobre la ventana principal.
        try:
            x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
            y = self.winfo_rooty() + 120
            win.geometry(f"+{max(0, x)}+{max(0, y)}")
        except Exception:
            pass

    def _actualizar_ahora(self):
        """Cierra y vuelve a abrir la app (el lanzador baja la versión nueva)."""
        if self.estado == "corriendo":
            if not messagebox.askyesno(
                    "Proceso en curso",
                    "Hay un registro en proceso. Si actualizas ahora se "
                    "interrumpirá.\n¿Continuar de todas formas?"):
                self.aviso_update_visible = False
                return
            self.detener_flag.set()
        try:
            import subprocess
            if getattr(sys, "frozen", False):
                subprocess.Popen([sys.executable])
            else:
                subprocess.Popen([sys.executable, os.path.abspath(sys.argv[0])])
        except Exception as e:
            messagebox.showerror("Error",
                                 f"No se pudo reiniciar automáticamente:\n{e}")
            self.aviso_update_visible = False
            return
        try:
            self.destroy()
        except Exception:
            pass
        os._exit(0)

    # ------------------------------------------------------------------ #
    #  Construcción de la interfaz
    # ------------------------------------------------------------------ #
    # ------------------------------------------------------------------ #
    #  Estilo / componentes visuales
    # ------------------------------------------------------------------ #
    def _estilo(self):
        """Configura el tema y los estilos ttk (look moderno claro)."""
        st = ttk.Style()
        try:
            st.theme_use("clam")
        except Exception:
            pass
        # Pestañas tipo "pasos".
        st.configure("UI.TNotebook", background=UI_BG, borderwidth=0,
                     tabmargins=(6, 6, 6, 0))
        st.configure("UI.TNotebook.Tab", font=(UI_FUENTE, 10, "bold"),
                     padding=(20, 9), background="#dbe2ea", foreground="#5b6b7b",
                     borderwidth=0)
        st.map("UI.TNotebook.Tab",
               background=[("selected", UI_CARD)],
               foreground=[("selected", UI_AZUL)],
               expand=[("selected", (0, 0, 0, 0))])
        # Tabla.
        st.configure("Treeview", rowheight=26, font=(UI_FUENTE, 9),
                     background=UI_CARD, fieldbackground=UI_CARD,
                     foreground=UI_TXT, borderwidth=0)
        st.configure("Treeview.Heading", font=(UI_FUENTE, 9, "bold"),
                     background="#eef2f6", foreground="#33475b",
                     relief="flat", padding=5)
        st.map("Treeview.Heading", background=[("active", "#e2e8ef")])
        st.map("Treeview", background=[("selected", "#cfe3ff")],
               foreground=[("selected", "#0a2540")])
        # Combobox.
        st.configure("TCombobox", padding=3)
        # Barra de progreso.
        st.configure("UI.Horizontal.TProgressbar", troughcolor="#d3dae2",
                     background=UI_VERDE, borderwidth=0, thickness=14)

    def _tarjeta(self, parent, titulo=""):
        """Crea una 'tarjeta' (fondo blanco, borde sutil, título azul).
        Devuelve (borde, cuerpo): empaqueta 'borde' y agrega hijos a 'cuerpo'."""
        borde = tk.Frame(parent, bg=UI_BORDE)
        inner = tk.Frame(borde, bg=UI_CARD)
        inner.pack(fill="both", expand=True, padx=1, pady=1)
        if titulo:
            tk.Label(inner, text=titulo, bg=UI_CARD, fg=UI_AZUL,
                     font=(UI_FUENTE, 10, "bold")).pack(anchor="w", padx=12,
                                                        pady=(8, 2))
        cuerpo = tk.Frame(inner, bg=UI_CARD)
        cuerpo.pack(fill="both", expand=True, padx=12, pady=(2, 10))
        return borde, cuerpo

    _ESTILOS_BOTON = {
        "primario": (UI_AZUL, "#0a5599", "white"),
        "exito":    (UI_VERDE, "#218a40", "white"),
        "peligro":  (UI_ROJO, "#cc1111", "white"),
        "neutro":   ("#e9edf2", "#dce3ec", "#1f2d3d"),
    }

    def _boton(self, parent, texto, comando, tipo="neutro", **kw):
        """Botón plano con color por acción y efecto hover."""
        bg, hover, fg = self._ESTILOS_BOTON.get(tipo, self._ESTILOS_BOTON["neutro"])
        b = tk.Button(parent, text=texto, command=comando, bg=bg, fg=fg,
                      activebackground=hover, activeforeground=fg,
                      disabledforeground="#aab2bc", relief="flat", bd=0,
                      cursor="hand2", font=(UI_FUENTE, 9, "bold"),
                      padx=14, pady=6, **kw)
        b.bind("<Enter>",
               lambda e: b["state"] != "disabled" and b.config(bg=hover))
        b.bind("<Leave>",
               lambda e: b["state"] != "disabled" and b.config(bg=bg))
        return b

    def _construir(self):
        self._estilo()
        self.configure(bg=UI_BG)

        # --- Encabezado (banner azul: logo + título + ambiente) ---
        cab = tk.Frame(self, bg=UI_AZUL, height=58)
        cab.pack(fill="x")
        cab.pack_propagate(False)
        try:
            img = tk.PhotoImage(file=ruta_recurso(
                os.path.join("Imagenes", "Quetzaltic Texto Blanco .png")))
            factor = max(1, img.height() // 40)  # ~40 px de alto
            self.logo_img = img.subsample(factor, factor)
            tk.Label(cab, image=self.logo_img, bg=UI_AZUL
                     ).pack(side="left", padx=(16, 10))
        except Exception:
            pass
        tk.Label(cab, text="RPA Solicitudes de Pago — SIPP", bg=UI_AZUL,
                 fg="white", font=(UI_FUENTE, 15, "bold")).pack(side="left", pady=12)
        amb_bg = "#bb0000" if config.AMBIENTE == "PRODUCCION" else "#c07a00"
        tk.Label(cab, text=f"  {config.AMBIENTE}  ", bg=amb_bg, fg="white",
                 font=(UI_FUENTE, 9, "bold")).pack(side="right", padx=16)
        tk.Frame(self, bg="#002b54", height=3).pack(fill="x")

        # --- Barra de estado inferior (persistente: progreso + estado) ---
        pie = tk.Frame(self, bg="#e7ebf0")
        pie.pack(side="bottom", fill="x")
        tk.Label(pie, text="Quetzaltic Solutions  ·  Hecho por Angel Baldi",
                 bg="#e7ebf0", fg="#8a93a0", font=(UI_FUENTE, 8)
                 ).pack(side="left", padx=12, pady=3)
        self.lbl_estado = tk.Label(pie, text="Listo", bg="#e7ebf0", fg="#445566",
                                   font=(UI_FUENTE, 9), anchor="e")
        self.lbl_estado.pack(side="right", padx=12)
        self.barra = ttk.Progressbar(pie, mode="determinate", length=240,
                                     style="UI.Horizontal.TProgressbar")
        self.barra.pack(side="right", padx=8, pady=5)

        # --- Cuerpo único con barra de desplazamiento vertical ---
        cont = tk.Frame(self, bg=UI_BG)
        cont.pack(fill="both", expand=True)
        self._canvas = tk.Canvas(cont, bg=UI_BG, highlightthickness=0, borderwidth=0)
        vsb = ttk.Scrollbar(cont, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        cuerpo = tk.Frame(self._canvas, bg=UI_BG, padx=10, pady=8)
        win_id = self._canvas.create_window((0, 0), window=cuerpo, anchor="nw")
        self._canvas.bind(
            "<Configure>",
            lambda e: self._canvas.itemconfigure(win_id, width=e.width))
        cuerpo.bind(
            "<Configure>",
            lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self.bind_all("<MouseWheel>", self._rueda)

        # ===================== Archivos de origen ===================== #
        b, f1 = self._tarjeta(cuerpo, "1 · Archivos de origen")
        b.pack(fill="x", pady=(0, 8))
        fmt = tk.Frame(f1, bg=UI_CARD)
        fmt.grid(row=0, column=0, sticky="w", pady=4)
        tk.Label(fmt, text="Formato:", bg=UI_CARD, fg=UI_TXT).pack(side="left")
        self.var_formato = tk.StringVar(value="CSV")
        ttk.Combobox(fmt, textvariable=self.var_formato, values=["CSV", "Excel"],
                     width=7, state="readonly").pack(side="left", padx=6)
        self._boton(f1, "Cargar base de datos…", self.cargar_base, "neutro"
                    ).grid(row=0, column=1, padx=6, pady=4, sticky="w")
        self.lbl_csv = tk.Label(f1, text="(ningún archivo cargado)", bg=UI_CARD,
                                fg=UI_TENUE)
        self.lbl_csv.grid(row=0, column=2, padx=8, sticky="w")
        self._boton(f1, "Carpeta de Carátulas…", self.seleccionar_caratulas,
                    "neutro").grid(row=1, column=0, padx=6, pady=4, sticky="w")
        self._boton(f1, "Archivos PDF (carátulas)…",
                    self.seleccionar_archivos_caratulas, "neutro"
                    ).grid(row=1, column=1, padx=6, pady=4, sticky="w")
        self.lbl_car = tk.Label(f1, text="(no seleccionada — requerida)",
                                bg=UI_CARD, fg=UI_TENUE)
        self.lbl_car.grid(row=1, column=2, padx=8, sticky="w")
        self._boton(f1, "Carpeta de Vo.Bo.…", self.seleccionar_vobo, "neutro"
                    ).grid(row=2, column=0, padx=6, pady=4, sticky="w")
        self._boton(f1, "Archivos PDF (Vo.Bo.)…", self.seleccionar_archivos_vobo,
                    "neutro").grid(row=2, column=1, padx=6, pady=4, sticky="w")
        self.lbl_vobo = tk.Label(f1, text="(opcional — no todos lo tienen)",
                                 bg=UI_CARD, fg=UI_TENUE)
        self.lbl_vobo.grid(row=2, column=2, padx=8, sticky="w")

        b, f1b = self._tarjeta(cuerpo, "1-B · Extraer datos con OCR")
        b.pack(fill="x", pady=8)
        tk.Label(f1b, text="Analiza las carátulas (banco, CLABE, RFC) y los recibos "
                 "Vo.Bo. (monto, nombre y RFC del colaborador) para llenar la tabla.",
                 bg=UI_CARD, fg=UI_TENUE, justify="left", wraplength=560
                 ).pack(side="left", anchor="w")
        self.btn_ocr = self._boton(f1b, "🔍  Analizar carátulas (OCR)",
                                   self.analizar_ocr, "primario")
        self.btn_ocr.configure(state="disabled")
        self.btn_ocr.pack(side="right", padx=4, pady=2)

        # ===================== Datos del lote ===================== #
        b, f1c = self._tarjeta(
            cuerpo, "2 · Datos del lote  (se aplican a los registros con ✓)")
        b.pack(fill="x", pady=(0, 8))
        tk.Label(f1c, text="Empresa:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=0, column=0, sticky="e", padx=4, pady=3)
        self.var_empresa = tk.StringVar(value="ABASTECEDORA")
        ttk.Combobox(f1c, textvariable=self.var_empresa, width=15, state="readonly",
                     values=sorted(mapeos.EMPRESAS.keys())
                     ).grid(row=0, column=1, sticky="w", padx=4)
        tk.Label(f1c, text="Sucursal:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=0, column=2, sticky="e", padx=4)
        self.var_sucursal = tk.StringVar(value="CORPORATIVO")
        ttk.Combobox(f1c, textvariable=self.var_sucursal, width=15, state="readonly",
                     values=sorted(mapeos.SUCURSALES.keys())
                     ).grid(row=0, column=3, sticky="w", padx=4)
        tk.Label(f1c, text="Moneda:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=0, column=4, sticky="e", padx=4)
        self.var_moneda = tk.StringVar(value="PESOS (MXN)")
        ttk.Combobox(f1c, textvariable=self.var_moneda, width=14, state="readonly",
                     values=sorted(mapeos.MONEDAS.keys())
                     ).grid(row=0, column=5, sticky="w", padx=4)
        tk.Label(f1c, text="Descripción:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=1, column=0, sticky="e", padx=4, pady=3)
        self.var_descripcion = tk.StringVar(value="PAGO DE UTILIDADES EX COLABORADOR")
        tk.Entry(f1c, textvariable=self.var_descripcion, width=40
                 ).grid(row=1, column=1, columnspan=3, sticky="w", padx=4)
        self.var_incluir_nombre = tk.BooleanVar(value=True)
        tk.Checkbutton(f1c, text="Incluir nombre del empleado", bg=UI_CARD,
                       activebackground=UI_CARD, fg=UI_TXT,
                       variable=self.var_incluir_nombre
                       ).grid(row=1, column=4, columnspan=2, sticky="w", padx=4)
        tk.Label(f1c, text="Fecha de pago:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=2, column=0, sticky="e", padx=4, pady=3)
        self.var_fecha = tk.StringVar(value="")
        tk.Entry(f1c, textvariable=self.var_fecha, width=15
                 ).grid(row=2, column=1, sticky="w", padx=4)
        tk.Label(f1c, text="(dd/mm/aaaa)", bg=UI_CARD, fg=UI_TENUE
                 ).grid(row=2, column=2, sticky="w")
        self._boton(f1c, "⬇  Aplicar a seleccionados",
                    self.aplicar_a_seleccionados, "exito"
                    ).grid(row=2, column=3, columnspan=3, padx=8, pady=4, sticky="w")

        # --- Vista previa (tabla) ---
        b, f2 = self._tarjeta(
            cuerpo, "Vista previa  (doble clic en Colaborador o Monto para editar)")
        b.pack(fill="both", expand=True, pady=8)
        self.lbl_resumen = tk.Label(
            f2, text="Carga un CSV o usa el OCR para ver el resumen.",
            bg=UI_CARD, fg=UI_TXT, anchor="w", justify="left")
        self.lbl_resumen.pack(fill="x")
        # Barra: seleccionar todos / conteo / descargar a Excel.
        barra_sel = tk.Frame(f2, bg=UI_CARD)
        barra_sel.pack(fill="x", side="bottom", pady=(6, 0))
        self.var_todos = tk.BooleanVar(value=True)
        tk.Checkbutton(barra_sel, text="Seleccionar todos", bg=UI_CARD,
                       activebackground=UI_CARD, fg=UI_TXT,
                       variable=self.var_todos, command=self._toggle_todos
                       ).pack(side="left", pady=2)
        self.lbl_sel = tk.Label(barra_sel, bg=UI_CARD, fg=UI_TENUE,
                                text="(clic en la casilla ✓ de cada renglón "
                                     "para incluir/excluir del proceso)")
        self.lbl_sel.pack(side="left", padx=8)
        self.btn_xlsx = self._boton(barra_sel, "📥  Descargar resumen (Excel)",
                                    self.exportar_xlsx, "exito")
        self.btn_xlsx.pack(side="right", padx=2, pady=2)
        # Contenedor de la tabla + scrollbars.
        tabla_wrap = tk.Frame(f2, bg=UI_CARD)
        tabla_wrap.pack(fill="both", expand=True, pady=(6, 0))
        cols = ("sel", "empresa", "sucursal", "codigo", "nombre", "rfc",
                "clabe", "banco", "valid18", "transf", "monto",
                "descripcion", "fecha", "obs")
        self.tabla = ttk.Treeview(tabla_wrap, columns=cols, show="headings",
                                  height=12)
        for c, t, w in [("sel", "✓", 30), ("empresa", "Empresa", 90),
                        ("sucursal", "Sucursal", 80), ("codigo", "Cód. colab.", 75),
                        ("nombre", "Nombre del colaborador", 175),
                        ("rfc", "RFC", 110), ("clabe", "CLABE", 145),
                        ("banco", "Banco", 85), ("valid18", "CLABE 18", 60),
                        ("transf", "Transf.", 55), ("monto", "Monto", 90),
                        ("descripcion", "Descripción", 150),
                        ("fecha", "Fecha pago", 85),
                        ("obs", "Observaciones (pendiente)", 230)]:
            self.tabla.heading(c, text=t)
            self.tabla.column(
                c, width=w, minwidth=40, stretch=(c == "obs"),
                anchor=("center" if c in ("sel", "valid18", "transf") else "w"))
        # Filas alternadas + estados (color del texto).
        self.tabla.tag_configure("par", background="#ffffff")
        self.tabla.tag_configure("impar", background="#f3f6f9")
        self.tabla.tag_configure("mal", foreground=COLOR_MAL)
        self.tabla.tag_configure("omit", foreground="#8a93a0")
        self.tabla.tag_configure("hecho", foreground=COLOR_OK)   # ya procesado
        sb = ttk.Scrollbar(tabla_wrap, orient="vertical", command=self.tabla.yview)
        hsb = ttk.Scrollbar(tabla_wrap, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(yscrollcommand=sb.set, xscrollcommand=hsb.set)
        sb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tabla.pack(side="left", fill="both", expand=True)
        self.tabla.bind("<Button-1>", self._click_tabla)
        self.tabla.bind("<Double-1>", self._editar_celda)
        self.item_a_fila = {}    # iid del Treeview -> fila
        self.fila_check = {}     # id(fila) -> bool (incluido en el proceso)
        self._editor = None      # Entry para editar celdas en línea

        # ================= TAB 3 — Ejecutar ================= #
        b, f3 = self._tarjeta(cuerpo, "3 · Credenciales SIPP")
        b.pack(fill="x", pady=(0, 8))
        tk.Label(f3, text="Usuario:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=0, column=0, sticky="e", padx=4, pady=4)
        self.ent_usuario = tk.Entry(f3, width=24)
        self.ent_usuario.grid(row=0, column=1, padx=4, pady=4)
        tk.Label(f3, text="Contraseña:", bg=UI_CARD, fg=UI_TXT
                 ).grid(row=0, column=2, sticky="e", padx=4, pady=4)
        self.ent_pass = tk.Entry(f3, width=24, show="*")
        self.ent_pass.grid(row=0, column=3, padx=4, pady=4)
        self.var_visible = tk.BooleanVar(value=True)
        tk.Checkbutton(f3, text="Mostrar navegador", bg=UI_CARD,
                       activebackground=UI_CARD, fg=UI_TXT,
                       variable=self.var_visible).grid(row=0, column=4, padx=12)
        # Mantener en estatus GUARDADO: no envía a autorizar (para validar antes).
        self.var_solo_guardar = tk.BooleanVar(value=False)
        tk.Checkbutton(
            f3, fg=UI_AZUL, bg=UI_CARD, activebackground=UI_CARD,
            text="Mantener en estatus guardado (NO enviar a autorizar)",
            variable=self.var_solo_guardar
        ).grid(row=1, column=0, columnspan=5, padx=4, pady=(0, 4), sticky="w")

        b, f4 = self._tarjeta(cuerpo, "4 · Ejecución")
        b.pack(fill="x", pady=8)
        self.btn_iniciar = self._boton(f4, "▶  Iniciar", self.iniciar, "primario")
        self.btn_iniciar.configure(state="disabled")
        self.btn_iniciar.pack(side="left", padx=4)
        self.btn_detener = self._boton(f4, "■  Detener", self.detener, "peligro")
        self.btn_detener.configure(state="disabled")
        self.btn_detener.pack(side="left", padx=4)
        self.btn_reanudar = self._boton(f4, "⏵  Reanudar", self.reanudar, "neutro")
        self.btn_reanudar.configure(state="disabled")
        self.btn_reanudar.pack(side="left", padx=4)
        self.btn_reiniciar = self._boton(f4, "↺  Reiniciar carga", self.reiniciar,
                                         "neutro")
        self.btn_reiniciar.configure(state="disabled")
        self.btn_reiniciar.pack(side="left", padx=4)
        self.btn_reporte = self._boton(f4, "📄  Generar reporte",
                                       self.generar_reporte, "neutro")
        self.btn_reporte.configure(state="disabled")
        self.btn_reporte.pack(side="left", padx=4)

        # --- Bitácora (avance) ---
        b, f5 = self._tarjeta(cuerpo, "Bitácora (avance)")
        b.pack(fill="both", expand=True, pady=8)
        self.txt = tk.Text(f5, height=12, state="disabled", wrap="word",
                           font=("Consolas", 9), bg="#fbfcfd", relief="flat",
                           bd=0)
        sb2 = ttk.Scrollbar(f5, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb2.set)
        self.txt.pack(side="left", fill="both", expand=True)
        sb2.pack(side="right", fill="y")

    def _enganchar_log(self):
        h = HandlerCola(self.cola)
        h.setFormatter(logging.Formatter("%(asctime)s | %(message)s", "%H:%M:%S"))
        logging.getLogger("rpa").addHandler(h)

    # ------------------------------------------------------------------ #
    #  Cargar y validar la base de datos (CSV o Excel)
    # ------------------------------------------------------------------ #
    def cargar_base(self):
        es_excel = self.var_formato.get() == "Excel"
        if es_excel:
            filtros = [("Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")]
            titulo = "Selecciona el archivo de Excel"
        else:
            filtros = [("CSV", "*.csv"), ("Todos", "*.*")]
            titulo = "Selecciona el archivo CSV"
        ruta = filedialog.askopenfilename(title=titulo, filetypes=filtros)
        if not ruta:
            return
        carpeta = os.path.dirname(ruta)
        config.ARCHIVO_CSV = ruta
        config.CARPETA_LOGS = os.path.join(carpeta, "logs")
        try:
            self.filas = sipp_rpa.leer_base(ruta)
        except Exception as e:
            messagebox.showerror("Error al leer la base de datos", str(e))
            return
        if not self.filas:
            messagebox.showwarning("Sin datos",
                                   "El archivo no tiene registros legibles.")
            return
        self.ruta_csv = ruta
        self.lbl_csv.config(text=os.path.basename(ruta), fg="#000")

        # Autodetecta carpetas junto al CSV (el operador puede cambiarlas).
        car = os.path.join(carpeta, "CARATULAS")
        vob = os.path.join(carpeta, "VOBO")
        if os.path.isdir(car):
            config.CARPETA_CARATULAS = car
            self._actualizar_lbl_archivos(self.lbl_car, car,
                                          config.ARCHIVOS_CARATULAS,
                                          "(no seleccionada — requerida)")
        if os.path.isdir(vob):
            config.CARPETA_VOBO = vob
            self._actualizar_lbl_archivos(self.lbl_vobo, vob,
                                          config.ARCHIVOS_VOBO,
                                          "(opcional — no todos lo tienen)")

        # Archivo nuevo: estado limpio (desde cero) y todo seleccionado.
        self._resetear_carga()
        self._mostrar_preview()
        self._actualizar_botones()

    def _resetear_carga(self):
        """Estado limpio para una carga nueva (CSV/Excel u OCR)."""
        self.estado = "idle"
        self.siguiente = 0
        self.resultados_global = []
        self.fila_check = {}
        self.procesados_ok = set()
        self.var_todos.set(True)

    # ------------------------------------------------------------------ #
    #  OCR: analiza las CARÁTULAS cargadas y llena la tabla
    # ------------------------------------------------------------------ #
    EXT_DOCS_OCR = (".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff",
                    ".heic", ".heif")

    def _archivos_caratula(self):
        """Carátulas cargadas (carpeta + archivos sueltos) que el OCR analizará."""
        rutas = list(config.ARCHIVOS_CARATULAS or [])
        carp = config.CARPETA_CARATULAS
        if carp and os.path.isdir(carp):
            for n in sorted(os.listdir(carp)):
                if n.lower().endswith(self.EXT_DOCS_OCR):
                    rutas.append(os.path.join(carp, n))
        return rutas

    def _archivos_vobo(self):
        """Recibos Vo.Bo. (carpeta + archivos sueltos) de donde el OCR saca el monto."""
        rutas = list(config.ARCHIVOS_VOBO or [])
        carp = config.CARPETA_VOBO
        if carp and os.path.isdir(carp):
            for n in sorted(os.listdir(carp)):
                if n.lower().endswith(self.EXT_DOCS_OCR):
                    rutas.append(os.path.join(carp, n))
        return rutas

    def analizar_ocr(self):
        if self.estado == "corriendo":
            return
        docs = self._archivos_caratula()
        if not docs:
            messagebox.showinfo(
                "OCR", "Primero carga las carátulas arriba (carpeta o archivos).")
            return
        vobos = self._archivos_vobo()
        self.btn_ocr.config(state="disabled")
        msg = f"=== OCR: analizando {len(docs)} carátula(s)"
        if vobos:
            msg += f" + {len(vobos)} recibo(s) Vo.Bo. para el monto"
        self._log_ui(msg + "... ===")
        threading.Thread(
            target=self._ocr_worker,
            args=(docs, vobos, self.var_empresa.get(), self.var_sucursal.get()),
            daemon=True).start()

    def _ocr_worker(self, docs, vobos, empresa, sucursal):
        try:
            import puente_ocr
            resultados = puente_ocr.extraer_documentos(
                docs,
                on_progreso=lambda i, t, n: self.cola.put(
                    ("log", f"   OCR [{i}/{t}]: {n}")))
            filas = [puente_ocr.resultado_a_fila(r, empresa=empresa,
                                                 sucursal=sucursal)
                     for r in resultados]
            # Lee los recibos Vo.Bo.: monto + nombre/código del colaborador.
            if vobos:
                self.cola.put(("log", "   OCR: leyendo recibos (Vo.Bo.) para monto y nombre..."))
                indice = puente_ocr.extraer_datos_vobo(
                    vobos,
                    on_progreso=lambda i, t, n: self.cola.put(
                        ("log", f"   Vo.Bo. [{i}/{t}]: {n}")))
                n_monto, n_nombre, n_rfc = puente_ocr.aplicar_datos_vobo(
                    filas, indice)
                self.cola.put(("log",
                    f"   OCR: {n_monto} monto(s), {n_nombre} nombre(s) y "
                    f"{n_rfc} RFC tomados de los recibos."))
            self.cola.put(("ocr_listo", filas))
        except Exception as e:
            self.cola.put(("ocr_error", str(e)))

    def _ocr_terminado(self, filas):
        self.btn_ocr.config(state="normal")
        if not filas:
            messagebox.showwarning("OCR", "No se extrajo ningún registro.")
            return
        self.filas = filas
        self.ruta_csv = None
        prim = filas[0].get("_CARATULA")     # logs junto a la 1ª carátula
        if prim:
            config.CARPETA_LOGS = os.path.join(os.path.dirname(prim), "logs")
        self.lbl_csv.config(text=f"{len(filas)} registro(s) extraídos por OCR",
                            fg=COLOR_OK)
        self._resetear_carga()
        self._mostrar_preview()
        self._actualizar_botones()
        rev = sum(1 for f in filas if f.get("_OCR_ESTADO") != "OK")
        self._log_ui(
            f"=== OCR listo: {len(filas)} registro(s). Revisa los datos y "
            f"agrega el MONTO (doble clic en la celda). "
            f"{rev} con observaciones del OCR. ===")

    def aplicar_a_seleccionados(self):
        """Aplica Empresa, Sucursal, Moneda, Descripción y Fecha del 'Paso 1-C'
        a los registros SELECCIONADOS (con ✓, menos los ya procesados). Si
        'Incluir nombre del empleado' está marcado, agrega el nombre al final de
        la descripción (útil para pago de utilidades de ex-colaboradores)."""
        if not self.filas:
            messagebox.showinfo("Aplicar a seleccionados",
                                "Primero carga registros (CSV/Excel u OCR).")
            return
        N = sipp_rpa.normalizar
        empresa = self.var_empresa.get()
        sucursal = self.var_sucursal.get()
        moneda = self.var_moneda.get()
        desc = self.var_descripcion.get().strip()
        incluir = self.var_incluir_nombre.get()
        fecha = self.var_fecha.get().strip()
        n = 0
        for f in self.filas:
            if self._esta_hecho(f):
                continue
            if not self.fila_check.get(id(f), True):
                continue   # sin palomita: no se le aplican los datos del lote
            if empresa:
                f[N("EMPRESA")] = empresa
            if sucursal:
                f[N("SUCURSAL")] = sucursal
            if moneda:
                f[N("MONEDA")] = moneda
            if fecha:
                f[N("FECHA DE PAGO")] = fecha
            if desc:
                if incluir:
                    nombre = sipp_rpa.campo(
                        f, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA").strip()
                    f[N("DESCRIPCION")] = f"{desc} - {nombre}" if nombre else desc
                else:
                    f[N("DESCRIPCION")] = desc
            n += 1
        if n == 0:
            messagebox.showinfo(
                "Aplicar a seleccionados",
                "No hay registros seleccionados (con ✓) para aplicarles los "
                "datos del lote.")
            return
        self._mostrar_preview()
        self._actualizar_botones()
        self._log_ui(f"Datos del lote aplicados a {n} registro(s) seleccionado(s).")

    FILTROS_PDF = [("PDF / imágenes", "*.pdf *.jpg *.jpeg *.png"),
                   ("Todos los archivos", "*.*")]

    def _lbl_carpeta(self, lbl, carpeta):
        """(Compatibilidad) Muestra solo la carpeta."""
        n = 0
        if os.path.isdir(carpeta):
            n = sum(1 for f in os.listdir(carpeta)
                    if os.path.splitext(f)[1].lower() in config.EXT_CARATULA)
        lbl.config(text=f"{carpeta}  ({n} archivos)", fg="#000")

    def _actualizar_lbl_archivos(self, lbl, carpeta, archivos, vacio_txt):
        """Muestra el estado combinado: carpeta y/o N archivos PDF sueltos."""
        partes = []
        if carpeta and os.path.isdir(carpeta):
            n = sum(1 for f in os.listdir(carpeta)
                    if os.path.splitext(f)[1].lower() in config.EXT_CARATULA)
            nombre = os.path.basename(carpeta.rstrip("\\/")) or carpeta
            partes.append(f"Carpeta '{nombre}' ({n})")
        if archivos:
            partes.append(f"{len(archivos)} archivo(s) PDF")
        if partes:
            lbl.config(text="  +  ".join(partes), fg=COLOR_OK)
        else:
            lbl.config(text=vacio_txt, fg="#555")

    def seleccionar_caratulas(self):
        d = filedialog.askdirectory(title="Carpeta de Carátulas")
        if not d:
            return
        config.CARPETA_CARATULAS = d
        self._actualizar_lbl_archivos(self.lbl_car, config.CARPETA_CARATULAS,
                                      config.ARCHIVOS_CARATULAS,
                                      "(no seleccionada — requerida)")
        self.btn_ocr.config(state="normal")   # ya hay carátulas para el OCR
        if self.filas:
            self._mostrar_preview()
        self._actualizar_botones()

    def seleccionar_archivos_caratulas(self):
        archivos = filedialog.askopenfilenames(
            title="Archivos de carátulas (PDF/imágenes)", filetypes=self.FILTROS_PDF)
        if not archivos:
            return
        config.ARCHIVOS_CARATULAS = list(archivos)
        self._actualizar_lbl_archivos(self.lbl_car, config.CARPETA_CARATULAS,
                                      config.ARCHIVOS_CARATULAS,
                                      "(no seleccionada — requerida)")
        self.btn_ocr.config(state="normal")   # ya hay carátulas para el OCR
        if self.filas:
            self._mostrar_preview()
        self._actualizar_botones()

    def seleccionar_vobo(self):
        d = filedialog.askdirectory(title="Carpeta de Vo.Bo.")
        if not d:
            return
        config.CARPETA_VOBO = d
        self._actualizar_lbl_archivos(self.lbl_vobo, config.CARPETA_VOBO,
                                      config.ARCHIVOS_VOBO,
                                      "(opcional — no todos lo tienen)")
        if self.filas:
            self._mostrar_preview()

    def seleccionar_archivos_vobo(self):
        archivos = filedialog.askopenfilenames(
            title="Archivos PDF de Vo.Bo.", filetypes=self.FILTROS_PDF)
        if not archivos:
            return
        config.ARCHIVOS_VOBO = list(archivos)
        self._actualizar_lbl_archivos(self.lbl_vobo, config.CARPETA_VOBO,
                                      config.ARCHIVOS_VOBO,
                                      "(opcional — no todos lo tienen)")
        if self.filas:
            self._mostrar_preview()

    def _mostrar_preview(self):
        for it in self.tabla.get_children():
            self.tabla.delete(it)
        self.item_a_fila = {}
        C = sipp_rpa.campo
        v = sipp_rpa.validar_datos(self.filas)
        probs = {n: e for n, e in v["problemas"]}
        for idx, fila in enumerate(self.filas):
            empresa = C(fila, "EMPRESA")
            sucursal = C(fila, "SUCURSAL")
            codigo = C(fila, "NUM. COLABORADOR", "CODIGO DE COLABORADOR", "CODIGO")
            nombre = C(fila, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA")
            rfc = C(fila, "RFC")
            clabe = C(fila, "CLAVE INTERBANCARIA", "CLABE")
            banco = C(fila, "BANCOS")
            transf = C(fila, "TIPO DE TRANSFERENCIA") or "SPEI"
            monto = sipp_rpa.campo_monto(fila).strip()
            descripcion = C(fila, "DESCRIPCION", "DESCRIPCIÓN SOLICITUD")
            fecha = C(fila, "FECHA DE PAGO", "FECHA")
            # Validación de 18 caracteres de la CLABE.
            ok18 = fila.get("_OCR_CLABE_OK")
            if ok18 is None:
                digs = "".join(ch for ch in clabe if ch.isdigit())
                ok18 = (len(digs) == 18)
            valid18 = "Sí" if ok18 else "No"
            # Estado / color de la fila + texto de Observaciones (lo pendiente).
            if sipp_rpa.clave_registro(fila) in self.procesados_ok:
                tags = ("hecho",)
                self.fila_check[id(fila)] = False
                glifo = "✔"
                obs = "Ya procesado ✓ (no se repite)"
            else:
                if not sipp_rpa.limpiar_monto(monto):
                    tags = ("omit",)
                    obs = "Sin monto (se omite)"
                elif nombre in probs:
                    tags = ("mal",)
                    obs = "; ".join(probs[nombre])
                else:
                    tags = ()
                    obs = "Listo"
                chk = self.fila_check.get(id(fila), True)
                self.fila_check[id(fila)] = chk
                glifo = "☑" if chk else "☐"
            # Fila alternada (fondo) + estado (color de texto).
            tags = (("par" if idx % 2 == 0 else "impar"),) + tags
            iid = self.tabla.insert(
                "", "end",
                values=(glifo, empresa, sucursal, codigo, nombre or "—", rfc,
                        clabe or "—", banco, valid18, transf, monto or "—",
                        descripcion, fecha, obs),
                tags=tags)
            self.item_a_fila[iid] = fila
        self._actualizar_conteo_sel()
        car = "OK" if v["hay_caratulas"] else "no detectada"
        nprob = len(v["problemas"])
        self.lbl_resumen.config(
            text=(f"Total: {v['total']}   |   Con monto: {v['con_monto']}   |   "
                  f"Sin monto (se omiten): {v['sin_monto']}   |   "
                  f"Con observaciones: {nprob}   |   Carátulas: {car}\n"
                  f"(verde = ya procesado · rojo = con observación · "
                  f"gris = sin monto · doble clic para editar)"),
            fg=(COLOR_MAL if nprob else COLOR_OK))

    def exportar_xlsx(self):
        """Descarga la vista previa (todos los registros, tal como se ven) a un
        archivo .xlsx, para tener una base de datos aparte."""
        if not self.filas:
            messagebox.showinfo("Descargar resumen",
                                "No hay registros para exportar.")
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar resumen como...",
            defaultextension=".xlsx",
            filetypes=[("Libro de Excel", "*.xlsx")],
            initialfile="Resumen solicitudes de pago.xlsx")
        if not ruta:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except Exception as e:
            messagebox.showerror(
                "Descargar resumen",
                f"No se pudo cargar openpyxl para crear el Excel:\n{e}")
            return
        C = sipp_rpa.campo
        v = sipp_rpa.validar_datos(self.filas)
        probs = {n: e for n, e in v["problemas"]}
        encabezados = ["Incluido", "Empresa", "Sucursal", "Cód. colaborador",
                       "Nombre del colaborador", "RFC", "CLABE", "Banco",
                       "CLABE 18 díg.", "Tipo transferencia", "Monto",
                       "Descripción", "Fecha de pago", "Observaciones"]
        anchos = [9, 16, 14, 14, 32, 15, 22, 16, 12, 16, 13, 34, 13, 34]
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes"
        ws.append(encabezados)
        hdr_fill = PatternFill("solid", fgColor="1A7F37")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for fila in self.filas:
            nombre = C(fila, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA")
            clabe = C(fila, "CLAVE INTERBANCARIA", "CLABE")
            monto_txt = sipp_rpa.campo_monto(fila).strip()
            limp = sipp_rpa.limpiar_monto(monto_txt)
            try:
                monto = float(limp) if limp else ""
            except ValueError:
                monto = monto_txt
            ok18 = fila.get("_OCR_CLABE_OK")
            if ok18 is None:
                digs = "".join(ch for ch in clabe if ch.isdigit())
                ok18 = (len(digs) == 18)
            if sipp_rpa.clave_registro(fila) in self.procesados_ok:
                obs = "Ya procesado"
            elif not sipp_rpa.limpiar_monto(monto_txt):
                obs = "Sin monto (se omite)"
            elif nombre in probs:
                obs = "; ".join(probs[nombre])
            else:
                obs = "Listo"
            incluido = "Sí" if self.fila_check.get(id(fila), True) else "No"
            ws.append([
                incluido,
                C(fila, "EMPRESA"),
                C(fila, "SUCURSAL"),
                C(fila, "NUM. COLABORADOR", "CODIGO DE COLABORADOR", "CODIGO"),
                nombre,
                C(fila, "RFC"),
                clabe,
                C(fila, "BANCOS"),
                "Sí" if ok18 else "No",
                C(fila, "TIPO DE TRANSFERENCIA") or "SPEI",
                monto,
                C(fila, "DESCRIPCION", "DESCRIPCIÓN SOLICITUD"),
                C(fila, "FECHA DE PAGO", "FECHA"),
                obs,
            ])
        # Formato: CLABE como texto, monto con separador de miles.
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=7).number_format = "@"          # CLABE
            ws.cell(row=r, column=11).number_format = '#,##0.00'  # Monto
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}1"
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        try:
            wb.save(ruta)
        except PermissionError:
            messagebox.showerror(
                "Descargar resumen",
                "No se pudo guardar (¿el archivo está abierto en Excel?).\n"
                "Ciérralo e inténtalo de nuevo.")
            return
        except Exception as e:
            messagebox.showerror("Descargar resumen", f"No se pudo guardar:\n{e}")
            return
        self._log_ui(f"Resumen exportado a Excel: {ruta}")
        if messagebox.askyesno(
                "Descargar resumen",
                f"Resumen guardado ({len(self.filas)} registro(s)) en:\n{ruta}\n\n"
                "¿Abrir el archivo ahora?"):
            try:
                os.startfile(ruta)
            except Exception:
                pass

    def _rueda(self, e):
        """Rueda del mouse: desplaza la ventana, salvo cuando el cursor está
        sobre la tabla o la bitácora (ahí scrollean ellas)."""
        w = self.winfo_containing(e.x_root, e.y_root)
        ruta = str(w) if w is not None else ""
        if ruta.startswith(str(self.tabla)) or ruta.startswith(str(self.txt)):
            return
        try:
            self._canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        except Exception:
            pass

    def _esta_hecho(self, fila):
        return sipp_rpa.clave_registro(fila) in self.procesados_ok

    def _click_tabla(self, event):
        """Clic en la casilla ✓ de un renglón -> incluye/excluye ese registro."""
        if self.tabla.identify_region(event.x, event.y) != "cell":
            return
        if self.tabla.identify_column(event.x) != "#1":   # solo la columna 'sel'
            return
        iid = self.tabla.identify_row(event.y)
        fila = self.item_a_fila.get(iid)
        if fila is None:
            return
        if self._esta_hecho(fila):
            return "break"   # ya procesado: no se puede volver a seleccionar
        nuevo = not self.fila_check.get(id(fila), True)
        self.fila_check[id(fila)] = nuevo
        self.tabla.set(iid, "sel", "☑" if nuevo else "☐")
        self._actualizar_conteo_sel()
        return "break"

    def _toggle_todos(self):
        """Casilla 'Seleccionar todos': marca/desmarca todos (menos los ya
        procesados, que quedan excluidos)."""
        val = self.var_todos.get()
        for iid, fila in self.item_a_fila.items():
            if self._esta_hecho(fila):
                continue
            self.fila_check[id(fila)] = val
            self.tabla.set(iid, "sel", "☑" if val else "☐")
        self._actualizar_conteo_sel()

    def _actualizar_conteo_sel(self):
        if not self.filas:
            self.lbl_sel.config(text="")
            return
        sel = sum(1 for f in self.filas if self.fila_check.get(id(f), True)
                  and not self._esta_hecho(f))
        hechos = sum(1 for f in self.filas if self._esta_hecho(f))
        extra = f"   |   Ya procesados: {hechos}" if hechos else ""
        self.lbl_sel.config(
            text=f"Seleccionados para procesar: {sel} de {len(self.filas)}{extra}")

    # ------------------------------------------------------------------ #
    #  Edición en línea (doble clic) de las celdas editables
    # ------------------------------------------------------------------ #
    # Columna del Treeview -> clave(s) de la fila que se actualizan.
    EDIT_MAP = {
        "empresa": ["EMPRESA"],
        "sucursal": ["SUCURSAL"],
        "codigo": ["NUM. COLABORADOR"],
        "nombre": ["EX-COLABORADOR (DESCRIPCION)", "NOMBRE DE CUENTA"],
        "rfc": ["RFC"],
        "clabe": ["CLAVE INTERBANCARIA"],
        "banco": ["BANCOS"],
        "monto": None,                       # especial: usa _set_monto
        "descripcion": ["DESCRIPCION"],
        "fecha": ["FECHA DE PAGO"],
    }

    def _editar_celda(self, event):
        if self.estado == "corriendo":
            return
        if self._editor is not None:
            try:
                self._editor.destroy()
            except Exception:
                pass
            self._editor = None
        if self.tabla.identify_region(event.x, event.y) != "cell":
            return
        col = self.tabla.identify_column(event.x)     # "#N"
        try:
            colname = self.tabla["columns"][int(col[1:]) - 1]
        except Exception:
            return
        if colname not in self.EDIT_MAP:              # columna no editable
            return
        iid = self.tabla.identify_row(event.y)
        fila = self.item_a_fila.get(iid)
        if fila is None or self._esta_hecho(fila):
            return
        celda = self.tabla.bbox(iid, col)
        if not celda:
            return
        x, y, w, h = celda
        valor = self.tabla.set(iid, colname).replace("—", "").strip()
        ent = tk.Entry(self.tabla)
        ent.place(x=x, y=y, width=max(w, 90), height=h)
        ent.insert(0, valor)
        ent.focus_set()
        ent.select_range(0, "end")
        self._editor = ent

        def commit(_=None):
            nuevo = ent.get().strip()
            if colname == "monto":
                self._set_monto(fila, nuevo)
            else:
                for k in self.EDIT_MAP[colname]:
                    fila[sipp_rpa.normalizar(k)] = nuevo
            try:
                ent.destroy()
            except Exception:
                pass
            self._editor = None
            self._mostrar_preview()

        def cancelar(_=None):
            try:
                ent.destroy()
            except Exception:
                pass
            self._editor = None

        ent.bind("<Return>", commit)
        ent.bind("<FocusOut>", commit)
        ent.bind("<Escape>", cancelar)

    def _set_monto(self, fila, valor):
        for k in list(fila.keys()):
            if "MONTO" in k or "IMPORTE" in k:
                fila[k] = valor
                return
        fila[sipp_rpa.normalizar("MONTO A PAGAR")] = valor

    # ------------------------------------------------------------------ #
    #  Ejecutar el robot
    # ------------------------------------------------------------------ #
    def _actualizar_botones(self, *_):
        # Hay carátula si: carpeta/archivos configurados, o cada fila trae su
        # propio documento (flujo OCR: fila['_CARATULA']).
        tiene_caratula = (os.path.isdir(config.CARPETA_CARATULAS)
                          or bool(config.ARCHIVOS_CARATULAS)
                          or bool(self.filas) and any(
                              f.get("_CARATULA") for f in self.filas))
        listo = (self.filas is not None and tiene_caratula)
        e = self.estado
        self.btn_iniciar.config(
            state="normal" if (listo and e in ("idle", "fin")) else "disabled")
        self.btn_detener.config(
            state="normal" if e == "corriendo" else "disabled")
        self.btn_reanudar.config(
            state="normal" if e == "detenido" else "disabled")
        self.btn_reiniciar.config(
            state="normal" if (listo and e in ("detenido", "fin")) else "disabled")
        # 'Generar reporte' disponible mientras NO esté corriendo (al detener,
        # al terminar, o tras cargar). Si no hay datos, el botón lo avisa.
        self.btn_reporte.config(
            state="normal" if (self.filas is not None and e != "corriendo")
            else "disabled")

    def generar_reporte(self):
        if not self.resultados_global:
            messagebox.showinfo(
                "Sin datos",
                "Todavía no hay registros procesados, así que no se generó "
                "ningún reporte.")
            return
        rep = sipp_rpa.escribir_reporte(self.resultados_global)
        if not rep:
            messagebox.showerror("Error", "No se pudo generar el reporte.")
            return
        messagebox.showinfo(
            "Reporte generado",
            f"Se generó el reporte con {len(self.resultados_global)} "
            f"registro(s):\n\n{rep}")
        if os.path.exists(rep) and messagebox.askyesno("Reporte", "¿Deseas abrirlo?"):
            try:
                os.startfile(rep)
            except Exception:
                pass

    def _credenciales(self):
        u = self.ent_usuario.get().strip()
        p = self.ent_pass.get()
        if not u or not p:
            messagebox.showwarning("Faltan credenciales",
                                   "Escribe tu usuario y contraseña de SIPP.")
            return None
        return u, p

    def _seleccionados(self):
        """Filas marcadas con la casilla ✓ y NO procesadas aún (evita duplicar)."""
        return [f for f in (self.filas or [])
                if self.fila_check.get(id(f), True) and not self._esta_hecho(f)]

    def _confirmar_inicio(self, seleccion):
        con_monto = sum(1 for f in seleccion
                        if sipp_rpa.limpiar_monto(sipp_rpa.campo_monto(f)))
        faltan = [sipp_rpa.campo(f, "EX-COLABORADOR (DESCRIPCIÓN)", "NOMBRE DE CUENTA")
                  for f in seleccion
                  if sipp_rpa.limpiar_monto(sipp_rpa.campo_monto(f))
                  and not sipp_rpa.caratula_de_fila(f)]
        if faltan and not messagebox.askyesno(
                "Faltan carátulas",
                f"{len(faltan)} registro(s) con monto NO tienen carátula y "
                f"fallarán al guardar.\nEjemplo: {faltan[0]}\n\n"
                f"¿Continuar de todas formas?"):
            return False
        if self.var_solo_guardar.get():
            modo = ("\n\nModo: se GUARDAN en estatus 'guardado' "
                    "(NO se envían a autorizar).")
        else:
            modo = "\n\nModo: se guardan y se ENVÍAN a autorizar."
        return messagebox.askyesno(
            f"Confirmar ejecución en {config.AMBIENTE}",
            f"Se registrarán {con_monto} solicitudes de pago en "
            f"{config.AMBIENTE}.{modo}\n\n¿Deseas continuar?")

    def iniciar(self):
        cred = self._credenciales()
        if not cred:
            return
        seleccion = self._seleccionados()
        if not seleccion:
            messagebox.showwarning(
                "Sin selección",
                "Marca al menos un registro (casilla ✓) para procesar.")
            return
        if not self._confirmar_inicio(seleccion):
            return
        self.usuario, self.contrasena = cred
        self.filas_run = seleccion
        self.siguiente = 0
        self.resultados_global = []
        self._log_ui(f"=== Iniciando ({len(seleccion)} registro(s) "
                     f"seleccionados) ===")
        self._arrancar()

    def reanudar(self):
        cred = self._credenciales()
        if not cred:
            return
        self.usuario, self.contrasena = cred
        self._log_ui(f"=== Reanudando desde el registro {self.siguiente + 1} ===")
        self._arrancar()

    def reiniciar(self):
        cred = self._credenciales()
        if not cred:
            return
        seleccion = self._seleccionados()
        if not seleccion:
            messagebox.showwarning(
                "Sin selección",
                "Marca al menos un registro (casilla ✓) para procesar.")
            return
        if not messagebox.askyesno(
                "Reiniciar carga",
                "Se volverá a empezar desde el PRIMER registro.\n\n"
                "OJO: los registros ya guardados en una corrida previa NO se "
                "borran; si los vuelves a procesar podrían duplicarse.\n\n"
                "¿Continuar?"):
            return
        self.usuario, self.contrasena = cred
        self.filas_run = seleccion
        self.siguiente = 0
        self.resultados_global = []
        self._log_ui("=== Reiniciando carga desde el primer registro ===")
        self._arrancar()

    def _arrancar(self):
        config.NAVEGADOR_VISIBLE = self.var_visible.get()
        # Si el operador eligió "mantener en guardado", NO se envía a autorizar.
        config.SOLICITAR_AUTORIZACION = not self.var_solo_guardar.get()
        self.detener_flag.clear()
        self.estado = "corriendo"
        self._actualizar_botones()
        self.barra.config(maximum=max(1, len(self.filas_run)), value=self.siguiente)
        inicio = self.siguiente
        self.worker = threading.Thread(
            target=self._correr, args=(inicio,), daemon=True)
        self.worker.start()

    def _correr(self, inicio):
        try:
            resumen = sipp_rpa.procesar(
                self.usuario, self.contrasena, self.filas_run[inicio:],
                on_progreso=lambda d: self.cola.put(("prog", {**d, "offset": inicio})),
                detener=self.detener_flag.is_set,
                escribir_rep=False)
            self.cola.put(("fin", resumen))
        except Exception as e:
            self.cola.put(("error", str(e)))

    def detener(self):
        self.detener_flag.set()
        self.lbl_estado.config(text="Deteniendo...")

    # ------------------------------------------------------------------ #
    #  Comunicación hilo -> ventana (cola)
    # ------------------------------------------------------------------ #
    def _log_ui(self, msg):
        self.txt.config(state="normal")
        self.txt.insert("end", msg + "\n")
        self.txt.see("end")
        self.txt.config(state="disabled")

    def _drenar_cola(self):
        try:
            while True:
                tipo, dato = self.cola.get_nowait()
                if tipo == "log":
                    self._log_ui(dato)
                elif tipo == "prog":
                    # Acumula el resultado EN VIVO (para poder generar el reporte
                    # en cualquier momento, incluso si se detiene).
                    res = dato.get("resultado")
                    if res:
                        self.resultados_global.append(res)
                        self.siguiente = len(self.resultados_global)
                    absoluto = dato.get("offset", 0) + dato.get("i", 0)
                    self.barra.config(value=absoluto)
                    self.lbl_estado.config(
                        text=f"{absoluto}/{len(self.filas_run)} {dato.get('estado','')}")
                elif tipo == "fin":
                    self._terminar(dato)
                elif tipo == "error":
                    self._terminar(None, error=dato)
                elif tipo == "version":
                    self._on_version(dato)
                elif tipo == "ocr_listo":
                    self._ocr_terminado(dato)
                elif tipo == "ocr_error":
                    self.btn_ocr.config(state="normal")
                    messagebox.showerror("OCR", f"Error al analizar:\n{dato}")
        except queue.Empty:
            pass
        self.after(150, self._drenar_cola)

    def _terminar(self, resumen, error=None):
        if error:
            self.estado = "detenido" if self.siguiente > 0 else "idle"
            self._actualizar_botones()
            self.lbl_estado.config(text="Error")
            messagebox.showerror("Error", f"El proceso falló:\n{error}")
            return
        r = resumen
        # Los resultados ya se fueron acumulando EN VIVO (self.resultados_global);
        # self.siguiente = cuántos se han procesado en total.
        detenido = r.get("detenido", False) and self.siguiente < len(self.filas_run)
        self.estado = "detenido" if detenido else "fin"
        self._actualizar_botones()

        # Marca como YA PROCESADOS los registros que salieron OK, para que no se
        # puedan volver a seleccionar/procesar en esta carga (evita duplicados).
        for x in self.resultados_global:
            if x.get("estado") == "OK" and x.get("clave"):
                self.procesados_ok.add(x["clave"])
        self._mostrar_preview()   # refresca: los OK quedan marcados y bloqueados

        cont = {}
        for x in self.resultados_global:
            cont[x["estado"]] = cont.get(x["estado"], 0) + 1
        estado_txt = "Detenido" if detenido else "Terminado"
        self.lbl_estado.config(text=estado_txt)
        rep = sipp_rpa.escribir_reporte(self.resultados_global)
        msg = (f"{estado_txt}. Procesados {self.siguiente} de {len(self.filas_run)}.\n\n"
               f"Éxitos: {cont.get('OK', 0)}\n"
               f"Cancelados (sin concepto): {cont.get('CANCELADO', 0)}\n"
               f"Revisar: {cont.get('REVISAR', 0)}\n"
               f"Errores: {cont.get('ERROR', 0)}\n"
               f"Omitidos (sin monto): {cont.get('OMITIDO', 0)}")
        if detenido:
            msg += ("\n\nProceso detenido. Usa 'Reanudar' para continuar desde "
                    "donde quedó, o 'Reiniciar carga' para empezar de nuevo.")
        if rep:
            msg += f"\n\nReporte:\n{rep}"
        messagebox.showinfo("Resultado", msg)
        if rep and os.path.exists(rep) and not detenido and messagebox.askyesno(
                "Reporte", "¿Deseas abrir el reporte final?"):
            try:
                os.startfile(rep)
            except Exception:
                pass


if __name__ == "__main__":
    App().mainloop()
